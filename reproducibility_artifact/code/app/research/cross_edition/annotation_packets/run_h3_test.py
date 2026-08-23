"""
One-off driver: tests H3 (PREREGISTRATION.md section 7) - "the method's
correspondence accuracy exceeds B2's (text-only)" - against the same 240
sampled items and the same complete, adjudicated ground truth already
computed in run_final_metrics.py. Not part of the pipeline itself.

Confirmation criterion (section 7): the paired difference (method accuracy
minus B2 accuracy) is positive with a bootstrap 95% CI whose lower bound
also excludes zero.
"""
import csv
import json
import random
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[4]))  # backend/ on sys.path

from app.research.cross_edition.annotation import (  # noqa: E402
    load_completed_xlsx, _norm_answer, majority_vote,
)
from app.research.cross_edition.baseline_b2 import align_items_b2  # noqa: E402
from app.research.cross_edition.annotation_packets.sample_join import (  # noqa: E402
    build_index_join, verify_sample_identity, join_baseline,
)

BASE = Path(__file__).parent

ANNOTATOR_FILES = {
    "A": r"C:\Users\Faisal\Desktop\research paper\Annotator_A.xlsx",
    "B": r"C:\Users\Faisal\Desktop\research paper\Annotator_B.xlsx",
    "C": r"C:\Users\Faisal\Desktop\research paper\Annotator_C.xlsx",
    "D": r"C:\Users\Faisal\Desktop\research paper\Annotator_D.xlsx",
}
ADJUDICATION_FILE = r"C:\Users\Faisal\Desktop\research paper\Adjudication_43_items_completed.xlsx"

# pair title -> (slug for the CSV/JSON dirs, old_pdf, new_pdf)
SP = r"C:\Users\Faisal\Desktop\Hospital SOP's Research\corpus\protocols"
PAIRS = {
    "Tennessee 2017→2018": ("tennessee_2017_2018", f"{SP}\\tn_2017.pdf", f"{SP}\\tn_2018.pdf"),
    "Pennsylvania 2021→2023": ("pennsylvania_2021_2023", f"{SP}\\pa_2021_als.pdf", f"{SP}\\pa_2023_als.pdf"),
    "Connecticut 2022.1→2023.1": ("connecticut_v20221_v20231", f"{SP}\\ct_v20221.pdf", f"{SP}\\ct_v20231.pdf"),
    "Connecticut 2023.1→2024.1": ("connecticut_v20231_v20241", f"{SP}\\ct_v20231.pdf", f"{SP}\\ct_v20241.pdf"),
}


def load_adjudicated() -> dict[str, dict[str, str]]:
    from openpyxl import load_workbook
    wb = load_workbook(ADJUDICATION_FILE, data_only=True)
    ws = wb["Adjudicate these 43"]
    out: dict[str, dict[str, str]] = {p: {} for p in PAIRS}
    for r in range(2, ws.max_row + 1):
        pair = ws.cell(row=r, column=1).value
        sid = ws.cell(row=r, column=2).value
        final_corr = ws.cell(row=r, column=9).value
        if pair and sid:
            out[pair][sid] = final_corr
    return out


def build_ground_truth() -> dict[str, dict[str, str]]:
    raw = {label: load_completed_xlsx(path) for label, path in ANNOTATOR_FILES.items()}
    adjudicated = load_adjudicated()
    ground_truth: dict[str, dict[str, str]] = {}
    for pair in PAIRS:
        norm = {
            label: {sid: _norm_answer(v["correspondence"])
                    for sid, v in raw[label][pair].items()}
            for label in ANNOTATOR_FILES
        }
        mv = majority_vote([norm["A"], norm["B"], norm["C"], norm["D"]])
        gt = {}
        for sid, result in mv.items():
            gt[sid] = result["answer"] if result["answer"] is not None \
                else _norm_answer(adjudicated[pair][sid])
        ground_truth[pair] = gt
    return ground_truth


def _weighted_mean_diff(sample: list[tuple[int, int, float]]) -> float:
    wsum = sum(w for _, _, w in sample)
    if wsum == 0:
        return 0.0
    return sum((m - b) * w for m, b, w in sample) / wsum


def bootstrap_ci(triples: list[tuple[int, int, float]], n_boot: int = 10000,
                  seed: int = 20261017):
    """triples: list of (method_correct, b2_correct, sample_weight) as
    0/1/float. Returns RAW (unweighted mean over the sample) and WEIGHTED
    (population-reweighted per section 5.1, since the stratified design
    oversamples rare tiers) point estimates, each with its own bootstrap
    95% CI, resampling item indices with replacement and keeping each
    item's own weight fixed within a resample."""
    rng = random.Random(seed)
    n = len(triples)
    raw_point = sum(m - b for m, b, _ in triples) / n
    weighted_point = _weighted_mean_diff(triples)

    raw_diffs, weighted_diffs = [], []
    for _ in range(n_boot):
        sample = [triples[rng.randrange(n)] for _ in range(n)]
        raw_diffs.append(sum(m - b for m, b, _ in sample) / n)
        weighted_diffs.append(_weighted_mean_diff(sample))
    raw_diffs.sort()
    weighted_diffs.sort()
    lo_i, hi_i = int(0.025 * n_boot), int(0.975 * n_boot) - 1

    return {
        "n": n, "n_bootstrap": n_boot,
        "raw": {"point_estimate": round(raw_point, 4),
                "ci95_low": round(raw_diffs[lo_i], 4),
                "ci95_high": round(raw_diffs[hi_i], 4)},
        "weighted": {"point_estimate": round(weighted_point, 4),
                     "ci95_low": round(weighted_diffs[lo_i], 4),
                     "ci95_high": round(weighted_diffs[hi_i], 4)},
    }


def main():
    ground_truth = build_ground_truth()

    all_paired: list[tuple[int, int]] = []
    per_pair_report = {}

    for pair, (slug, old_pdf, new_pdf) in PAIRS.items():
        packet_csv = BASE / slug / "annotation_packet.csv"
        with open(packet_csv, encoding="utf-8") as f:
            method_rows = {r["sample_id"]: r for r in csv.DictReader(f)}

        # Joined by PARSE-ORDER INDEX, not by old_item_id string - see
        # sample_join.py's module docstring. align_items mutates item_id
        # into the new edition's guideline vocabulary; annotation_packet.csv
        # stores that post-remap id while B2 (via align_items_b2 -> parse())
        # keys its own results by the raw parse() id. The id-based lookup
        # this loop used before the 2026-08-18 audit silently dropped any
        # sampled item whose guideline title changed between editions - 24
        # of 233 usable items project-wide, 21 of them (36%) in Connecticut
        # #1 alone - non-randomly inflating the reported method accuracy.
        b2 = align_items_b2(old_pdf, new_pdf)
        id_to_index, _ = build_index_join(old_pdf, new_pdf)
        verify_sample_identity(packet_csv, id_to_index)  # fail loudly, not silently

        gt = ground_truth[pair]
        pair_triples: list[tuple[int, int, float]] = []
        skipped_cannot_determine = 0
        for sid, row in method_rows.items():
            if sid not in gt:
                continue
            truth = gt[sid]
            if truth == "cannot_determine":
                skipped_cannot_determine += 1
                continue
            method_answer = _norm_answer(row["method_predicted_item_id"])
            idx = id_to_index[row["old_item_id"]]  # KeyError, not silent skip, if this ever fails
            b2_answer = _norm_answer(join_baseline(b2, idx, "b2_predicted_item_id"))
            weight = float(row.get("sample_weight", 1.0) or 1.0)

            m_correct = 1 if method_answer == truth else 0
            b_correct = 1 if b2_answer == truth else 0
            pair_triples.append((m_correct, b_correct, weight))
            all_paired.append((m_correct, b_correct, weight))

        method_acc = sum(m for m, _, _ in pair_triples) / len(pair_triples)
        b2_acc = sum(b for _, b, _ in pair_triples) / len(pair_triples)
        boot = bootstrap_ci(pair_triples)
        per_pair_report[pair] = {
            "n": len(pair_triples),
            "n_cannot_determine_excluded": skipped_cannot_determine,
            "method_accuracy": round(method_acc, 4),
            "b2_accuracy": round(b2_acc, 4),
            "bootstrap": boot,
        }
        print(f"=== {pair} ===")
        print(f"  method accuracy: {method_acc:.4f}   B2 accuracy: {b2_acc:.4f}")
        print(f"  paired diff raw (95% CI):      {boot['raw']}")
        print(f"  paired diff weighted (95% CI): {boot['weighted']}")
        print()

    pooled_method_acc = sum(m for m, _, _ in all_paired) / len(all_paired)
    pooled_b2_acc = sum(b for _, b, _ in all_paired) / len(all_paired)
    pooled_boot = bootstrap_ci(all_paired)
    confirmed_raw = pooled_boot["raw"]["point_estimate"] > 0 and pooled_boot["raw"]["ci95_low"] > 0
    confirmed_weighted = pooled_boot["weighted"]["point_estimate"] > 0 and pooled_boot["weighted"]["ci95_low"] > 0

    print("=== POOLED (all 4 pairs) ===")
    print(f"  method accuracy: {pooled_method_acc:.4f}   B2 accuracy: {pooled_b2_acc:.4f}")
    print(f"  paired diff raw (95% CI):      {pooled_boot['raw']}")
    print(f"  paired diff weighted (95% CI): {pooled_boot['weighted']}")
    print(f"  H3 (raw)      {'CONFIRMED' if confirmed_raw else 'NOT CONFIRMED'}")
    print(f"  H3 (weighted) {'CONFIRMED' if confirmed_weighted else 'NOT CONFIRMED'}")

    report = {
        "_superseded_by": (
            "full_comparison_report.json's 'H3 re-run' section, which "
            "recomputes this exact same analysis (method vs B2 accuracy, "
            "item-level and pair-level, raw and weighted) alongside H4/H5 "
            "with Benjamini-Hochberg correction across all three. This "
            "file is kept only as a standalone single-hypothesis check; "
            "do not cite it in place of full_comparison_report.json."
        ),
        "pairs": per_pair_report,
        "pooled": {
            "n": len(all_paired),
            "method_accuracy": round(pooled_method_acc, 4),
            "b2_accuracy": round(pooled_b2_acc, 4),
            "bootstrap": pooled_boot,
            "h3_confirmed_raw": confirmed_raw,
            "h3_confirmed_weighted": confirmed_weighted,
        },
    }
    out_path = BASE / "h3_test_report.SUPERSEDED.json"
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(report, f, indent=2, ensure_ascii=False)
    print(f"\nwrote {out_path}")


if __name__ == "__main__":
    main()
