"""
Builds the boundary-annotation workbooks for the novelty-audit's structure-
quality calibration (Workstream A). Annotators reconstruct an ORDERED LIST
of distinct protocol/guideline titles by reading each Tennessee edition's
PDF directly - the same "hand-built table of contents" approach that
originally diagnosed the section 56 boundary-detection bug - rather than
page/line numbers, since item_parser.py's canonical_text offsets do not
map cleanly back to physical PDF pages (extract_text joins all pages'
text with no page-boundary markers preserved).

Scope: Tennessee's four editions only (2017, 2018, 2022-23, Sept2024).
This is the one format where the section 56 bug is directly diagnosed and
confirmed, and all four editions share the identical known-anchor
detection code path, giving four independent, directly comparable
measurements. Pennsylvania (zero flagged outliers in section 58) and
Connecticut (boundaries come from its own Table of Contents via
detect_ct_toc_anchors - using annotator-read boundaries there would be
circularly validating the same source the parser already reads) are
deliberately excluded, per PREREGISTRATION.md's calibration entry.

Task design: annotators are told explicitly to record GUIDELINE-level
titles only (distinct named protocols - "Adult Cardiac Arrest",
"Anaphylaxis") and NOT the repeating sub-section headings that appear
inside many different protocols (Tennessee's own template slots,
empirically identified from corpus_probe output: "Treatment Pathway",
"Paramedic Stop", "Signs and Symptoms", "AEMT/EMT Stop Here",
"Paramedic", "Reference", "Paramedic Only", "Medical Emergency",
"Procedure") - this is exactly the distinction item_parser.py's own
detect_section_names function has to make automatically, so making it
explicit here is not an unfair simplification, it is the actual task.

Not part of the research pipeline itself.

Run:
    cd sop-guard/backend
    python -m app.research.cross_edition.annotation_packets.build_boundary_workbooks
"""
from __future__ import annotations

import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[4]))  # backend/ on sys.path

from openpyxl import Workbook  # noqa: E402
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side  # noqa: E402
from openpyxl.utils import get_column_letter  # noqa: E402

BASE = Path(__file__).parent / "boundary_annotation"

EDITIONS = [
    ("tn_2017", "Tennessee 2017 (Rev 11.7.2017)"),
    ("tn_2018", "Tennessee 2018 (Rev 7.7.18)"),
    ("tn_2022_23", "Tennessee 2022-23"),
    ("tn_sept2024", "Tennessee Sept2024 (2024-2025)"),
]

FONT_NAME = "Calibri"
HEADER_FILL = PatternFill("solid", fgColor="4472C4")
HEADER_FONT = Font(name=FONT_NAME, bold=True, color="FFFFFF", size=11)
FILLIN_FILL = PatternFill("solid", fgColor="FFF2CC")
NORMAL_FONT = Font(name=FONT_NAME, size=11)
WRAP_TOP = Alignment(wrap_text=True, vertical="top")
THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

KNOWN_SUBSECTIONS = (
    "Treatment Pathway, Paramedic Stop, Signs and Symptoms, AEMT Stop Here, "
    "EMT Stop Here, Paramedic, Reference, Paramedic Only, Medical Emergency, "
    "Procedure, Indications, Contraindications, Notes"
)

INSTRUCTIONS = [
    ("What you're doing", None),
    (None, "You're reading through one edition of a Tennessee EMS protocol "
           "PDF (open it alongside this workbook - one file per tab, same "
           "name) and writing down, IN ORDER, every distinct protocol's "
           "name as you reach it - like building a table of contents by "
           "hand, from the document itself, not from any existing index."),
    ("What counts as a protocol boundary", None),
    (None, "A GUIDELINE is a distinct named clinical protocol - e.g. "
           "'Adult Cardiac Arrest', 'Anaphylaxis', 'Torsades de Pointes'. "
           "Record its title exactly as printed, once, the first time it "
           "starts."),
    (None, f"Do NOT record these - they are recurring SUB-HEADINGS that "
           f"appear inside many different protocols, not protocols "
           f"themselves: {KNOWN_SUBSECTIONS}. If you see a heading not on "
           f"this list that also looks like a repeating template slot "
           f"rather than a distinct clinical topic, use your judgement and "
           f"note it in the Notes column."),
    ("How to work", None),
    (None, "Flip through the PDF page by page (skimming is fine - protocol "
           "titles are the larger/bolder headings). Each time you reach a "
           "new protocol, add one row: its title, and the page number "
           "you're on (approximate is fine, just for your own reference)."),
    ("Rules", None),
    (None, "1. Work alone. Don't discuss with the other annotator, or "
           "compare lists, until you have BOTH finished all four tabs."),
    (None, "2. Go in document order, front to back. Don't skip around."),
    (None, "3. If you're genuinely unsure whether something is a protocol "
           "title or a sub-heading, include it and note your uncertainty "
           "in the Notes column rather than guessing silently."),
    ("When you're done", None),
    (None, "Save this file and send it back exactly as-is."),
]

COLUMNS = [
    ("#", 6, "n"),
    ("Protocol/guideline title (as printed)", 55, "title"),
    ("Approx. page", 12, "page"),
    ("Notes (optional)", 30, "notes"),
]


def build_instructions_sheet(wb: Workbook, annotator_label: str):
    ws = wb.active
    ws.title = "READ ME FIRST"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 100
    r = 1
    c = ws.cell(row=r, column=1, value=f"Boundary annotation task - {annotator_label}")
    c.font = Font(name=FONT_NAME, bold=True, size=16, color="1F4E78")
    r += 2
    for heading, body in INSTRUCTIONS:
        if heading is not None:
            c = ws.cell(row=r, column=1, value=heading)
            c.font = Font(name=FONT_NAME, bold=True, size=13, color="1F4E78")
            r += 1
        if body is not None:
            c = ws.cell(row=r, column=1, value=body)
            c.font = Font(name=FONT_NAME, size=11.5)
            c.alignment = Alignment(wrap_text=True, vertical="top")
            ws.row_dimensions[r].height = 40
            r += 1
    r += 1
    note = ws.cell(row=r, column=1, value="Tabs: one per edition PDF. ~40-77 rows expected per tab (matches item_parser.py's own detected guideline counts).")
    note.font = Font(name=FONT_NAME, italic=True, size=11, color="595959")


def build_edition_sheet(wb: Workbook, sheet_title: str, n_prefill_rows: int = 100):
    ws = wb.create_sheet(title=sheet_title[:31])
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"
    for col_idx, (header, width, _key) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = WRAP_TOP
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[1].height = 24
    for i in range(n_prefill_rows):
        r = i + 2
        for col_idx, (_h, _w, key) in enumerate(COLUMNS, start=1):
            cell = ws.cell(row=r, column=col_idx, value=(i + 1) if key == "n" else "")
            cell.border = BORDER
            cell.alignment = WRAP_TOP
            cell.font = FILLIN_FONT if key != "n" else NORMAL_FONT
            if key in ("title", "page", "notes"):
                cell.fill = FILLIN_FILL
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{n_prefill_rows + 1}"


FILLIN_FONT = Font(name=FONT_NAME, size=11)


def build_workbook(annotator_label: str, out_path: Path):
    wb = Workbook()
    build_instructions_sheet(wb, annotator_label)
    for slug, title in EDITIONS:
        build_edition_sheet(wb, title)
    wb.save(out_path)
    print(f"wrote {out_path}")


if __name__ == "__main__":
    out_dir = BASE
    out_dir.mkdir(parents=True, exist_ok=True)
    build_workbook("Boundary Annotator 1", out_dir / "Boundary_Annotator_1.xlsx")
    build_workbook("Boundary Annotator 2", out_dir / "Boundary_Annotator_2.xlsx")
