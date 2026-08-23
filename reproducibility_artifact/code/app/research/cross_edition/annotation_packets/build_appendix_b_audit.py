"""
Appendix B item 3 audit workbook (audit round 4, Phase 4): "A manual
audit of all accepted guideline pairs is required before publication" -
registered at pre-registration, never done until now.

Rather than asking a human to review all ~130-190 accepted guideline
pairs per confirmatory edition pair, this generates a RANKED, PRE-
FLAGGED workbook so the reviewer sees the pairs most likely to exhibit
the two known failure modes first:

1. CONTAINMENT: match_guidelines' scoring is j = |overlap| / min(|a|,|b|)
   - a short title's tokens being a strict subset of a longer title's
     tokens scores a perfect 1.0, the exact mechanism that produced the
     documented "Hypothermia" -> "Induced Hypothermia Following ROSC"
     collision (PREREGISTRATION.md's 2026-08-18 boundary-annotation
     entry) and the containment bias Appendix B item 3 itself names.
2. TIES: `tiebreak_sensitivity.py` (audit round 3) measured 31/92 (34%)
   and 33/93 (35%) of Connecticut's old guidelines have multiple
   candidates scoring the SAME j=1.0, resolved only by an alphabetical
   tie-break with no semantic basis - any of those ties could be the
   wrong pick.

Reuses item_align.match_guidelines, _norm, and _TITLE_FLOOR completely
unchanged - this recomputes the SAME candidate scores match_guidelines
itself computes internally (which it doesn't expose), it does not
invent a new scoring method.

Does not modify item_align.py, item_parser.py, corpus_probe.py, or
edition_align.py.

Run:
    cd sop-guard/backend
    python -m app.research.cross_edition.annotation_packets.build_appendix_b_audit
"""
from __future__ import annotations

import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[4]))  # backend/ on sys.path

from openpyxl import Workbook  # noqa: E402
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side  # noqa: E402
from openpyxl.utils import get_column_letter  # noqa: E402

from app.research.cross_edition.item_parser import parse  # noqa: E402
from app.research.cross_edition.item_align import match_guidelines, _norm, _TITLE_FLOOR  # noqa: E402
from app.research.cross_edition.annotation_packets.run_h3_test import PAIRS  # noqa: E402

BASE = Path(__file__).parent

FONT_NAME = "Calibri"
HEADER_FONT = Font(name=FONT_NAME, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
FLAG_FILL = PatternFill("solid", fgColor="FFF2CC")
BORDER = Border(*[Side(style="thin", color="D9D9D9")] * 4)
WRAP_TOP = Alignment(wrap_text=True, vertical="top")


def _toks(t: str) -> set[str]:
    return {w for w in _norm(t).split() if len(w) > 2}


def score_all_candidates(old_titles: list[str], new_titles: list[str]) -> list[tuple]:
    """Reimplements match_guidelines' own scoring loop (identical formula,
    identical floor) to expose every candidate pair, not just the final
    greedy selection - match_guidelines itself does not return this."""
    pairs = []
    for o in old_titles:
        to = _toks(o)
        if not to:
            continue
        for n in new_titles:
            tn = _toks(n)
            if not tn:
                continue
            j = len(to & tn) / min(len(to), len(tn))
            if j >= _TITLE_FLOOR:
                pairs.append((j, o, n))
    return pairs


def audit_pair(pair_label: str, old_pdf: str, new_pdf: str) -> list[dict]:
    old_ed, new_ed = parse(old_pdf), parse(new_pdf)
    old_titles = sorted({i.guideline for i in old_ed.items})
    new_titles = sorted({i.guideline for i in new_ed.items})

    all_candidates = score_all_candidates(old_titles, new_titles)
    accepted = match_guidelines(old_titles, new_titles)

    # Group all candidates by old title for per-old-title analysis.
    by_old: dict[str, list[tuple]] = {}
    for j, o, n in all_candidates:
        by_old.setdefault(o, []).append((j, n))

    rows = []
    for o, accepted_new in accepted.items():
        candidates = sorted(by_old.get(o, []), reverse=True)
        accepted_score = next((j for j, n in candidates if n == accepted_new), None)
        if accepted_score is None:
            continue

        to, tn = _toks(o), _toks(accepted_new)
        is_containment = (
            accepted_score >= 0.999 and len(to) != len(tn)
            and (to <= tn or tn <= to)
        )
        n_tied = sum(1 for j, n in candidates if j >= accepted_score - 1e-9 and n != accepted_new)
        alt_candidates = [n for j, n in candidates
                           if n != accepted_new and j >= accepted_score - 0.15][:5]

        flag_reasons = []
        if is_containment:
            flag_reasons.append("CONTAINMENT (score=1.0, unequal length - Appendix B item 3)")
        if n_tied > 0:
            flag_reasons.append(f"TIE ({n_tied} other candidate(s) score >= accepted)")

        rows.append({
            "pair": pair_label, "old_title": o, "new_title": accepted_new,
            "score": round(accepted_score, 4), "flags": "; ".join(flag_reasons),
            "flag_priority": (2 if is_containment else 0) + (1 if n_tied > 0 else 0),
            "n_tied": n_tied, "alt_candidates": "; ".join(alt_candidates),
        })
    return rows


def build_workbook(all_rows: list[dict], out_path: Path):
    wb = Workbook()
    ws = wb.active
    ws.title = "READ ME FIRST"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 100
    instructions = [
        ("What this is", None),
        (None, "PREREGISTRATION.md's Appendix B item 3 states: 'Guideline matching is "
               "permissive by design - containment-biased overlap paired Cyanide "
               "Exposure with a truncated Exposure. A manual audit of all accepted "
               "guideline pairs is required before publication.' This workbook is "
               "that audit, generated but not yet reviewed."),
        ("How this is organized", None),
        (None, "The 'Flagged pairs' tab lists every accepted old-guideline-to-new-"
               "guideline match across the 4 confirmatory edition pairs, RANKED so "
               "the pairs most likely to be wrong appear first: CONTAINMENT flags "
               "(a short title's words are a strict subset of a longer title's - the "
               "exact mechanism that produced the documented Hypothermia -> Induced "
               "Hypothermia Following ROSC collision) and TIE flags (multiple "
               "candidate titles scored equally, so the accepted one was picked by "
               "an alphabetical tie-break with no semantic basis - measured to affect "
               "34-35% of Connecticut's old guidelines in a prior audit)."),
        ("What to do", None),
        (None, "For each flagged row, open both PDFs (or use the Score/Alt candidates "
               "columns) and judge: is the accepted new_title genuinely the same "
               "clinical protocol as old_title, or did it steal the match from a "
               "better candidate (listed in Alt candidates)? Mark your verdict in the "
               "'Verdict' column: CORRECT, WRONG, or UNSURE, with a note if useful."),
        ("Scope", None),
        (None, "You do not need to review every row - rows are sorted by "
               "flag_priority (highest first); the unflagged rows at the bottom "
               "(clean 1:1 matches, no ties, no containment) are lowest priority and "
               "included only for completeness."),
    ]
    r = 1
    c = ws.cell(row=r, column=1, value="Appendix B item 3: guideline-pair audit")
    c.font = Font(name=FONT_NAME, bold=True, size=16, color="1F4E78")
    r += 2
    for heading, body in instructions:
        if heading is not None:
            c = ws.cell(row=r, column=1, value=heading)
            c.font = Font(name=FONT_NAME, bold=True, size=13, color="1F4E78")
            r += 1
        if body is not None:
            c = ws.cell(row=r, column=1, value=body)
            c.font = Font(name=FONT_NAME, size=11.5)
            c.alignment = WRAP_TOP
            ws.row_dimensions[r].height = 55
            r += 1

    ws2 = wb.create_sheet("Flagged pairs")
    columns = [
        ("Pair", 24), ("Old title", 38), ("New title", 38), ("Score", 9),
        ("Flags", 40), ("Alt candidates (top 5)", 45), ("Verdict", 14), ("Notes", 30),
    ]
    for col_idx, (header, width) in enumerate(columns, start=1):
        cell = ws2.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = WRAP_TOP
        cell.border = BORDER
        ws2.column_dimensions[get_column_letter(col_idx)].width = width
    ws2.freeze_panes = "A2"

    sorted_rows = sorted(all_rows, key=lambda r: (-r["flag_priority"], r["pair"], r["old_title"]))
    for i, row in enumerate(sorted_rows, start=2):
        values = [row["pair"], row["old_title"], row["new_title"], row["score"],
                  row["flags"], row["alt_candidates"], "", ""]
        for col_idx, v in enumerate(values, start=1):
            cell = ws2.cell(row=i, column=col_idx, value=v)
            cell.border = BORDER
            cell.alignment = WRAP_TOP
            if row["flag_priority"] > 0:
                cell.fill = FLAG_FILL
    ws2.auto_filter.ref = f"A1:{get_column_letter(len(columns))}{len(sorted_rows) + 1}"

    n_flagged = sum(1 for r in all_rows if r["flag_priority"] > 0)
    n_containment = sum(1 for r in all_rows if "CONTAINMENT" in r["flags"])
    n_tied = sum(1 for r in all_rows if "TIE" in r["flags"])
    print(f"Total accepted pairs: {len(all_rows)}")
    print(f"Flagged: {n_flagged} ({n_containment} containment, {n_tied} tied)")

    wb.save(out_path)
    print(f"wrote {out_path}")


def main():
    all_rows = []
    for pair, (slug, old_pdf, new_pdf) in PAIRS.items():
        print(f"Auditing {pair}...")
        rows = audit_pair(pair, old_pdf, new_pdf)
        all_rows.extend(rows)
        n_flagged = sum(1 for r in rows if r["flag_priority"] > 0)
        print(f"  {len(rows)} accepted pairs, {n_flagged} flagged")

    out_path = BASE / "Appendix_B_guideline_pair_audit.xlsx"
    build_workbook(all_rows, out_path)


if __name__ == "__main__":
    main()
