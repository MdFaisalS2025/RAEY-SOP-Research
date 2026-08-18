"""
One-off script: builds a single adjudication workbook covering the 43
items (across all 4 pairs) that had no majority answer among the four
annotators (see 4rater_analysis_report.json, PREREGISTRATION.md section 11
2026-08-17 "Annotation upgraded to four annotators").

Not part of the pipeline itself - a deliverable-formatting step over
already-collected annotation data, same category as
build_annotator_workbooks.py.
"""
import csv
import json
import re
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

BASE = Path(__file__).parent
sys.path.insert(0, str(BASE.parents[3]))  # backend/ on sys.path
from app.research.cross_edition.annotation import load_completed_xlsx  # noqa: E402

_ILLEGAL_XML_CHARS_RE = re.compile(
    "[\x00-\x08\x0b\x0c\x0e-\x1f\x7f-\x84\x86-\x9f\ud800-\udfff﷐-﷯￾￿]"
)


def sanitize(text) -> str:
    if not isinstance(text, str):
        return text
    return _ILLEGAL_XML_CHARS_RE.sub("", text)


PAIR_SLUGS = {
    "Tennessee 2017→2018": "tennessee_2017_2018",
    "Pennsylvania 2021→2023": "pennsylvania_2021_2023",
    "Connecticut 2022.1→2023.1": "connecticut_v20221_v20231",
    "Connecticut 2023.1→2024.1": "connecticut_v20231_v20241",
}

ANNOTATOR_FILES = {
    "A": r"C:\Users\Faisal\Desktop\Annotator_A.xlsx",
    "B": r"C:\Users\Faisal\Desktop\Annotator_B.xlsx",
    "C": r"C:\Users\Faisal\Desktop\Annotator_C.xlsx",
    "D": r"C:\Users\Faisal\Desktop\Annotator_D.xlsx",
}

FONT_NAME = "Calibri"
HEADER_FILL = PatternFill("solid", fgColor="C00000")
HEADER_FONT = Font(name=FONT_NAME, bold=True, color="FFFFFF", size=11)
VOTE_FILL = PatternFill("solid", fgColor="F2F2F2")
FILLIN_FILL = PatternFill("solid", fgColor="FFF2CC")
NORMAL_FONT = Font(name=FONT_NAME, size=11)
WRAP_TOP = Alignment(wrap_text=True, vertical="top")
THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
RELATION_OPTIONS = "unchanged,reworded,substantive,merged,split,moved"

COLUMNS = [
    ("Pair", 20, "pair"),
    ("sample_id", 9, "sample_id"),
    ("OLD RECOMMENDATION", 40, "old_text"),
    ("NEW EDITION - full matching guideline", 50, "new_guideline_text"),
    ("Annotator A said", 30, "vote_A"),
    ("Annotator B said", 30, "vote_B"),
    ("Annotator C said", 30, "vote_C"),
    ("Annotator D said", 30, "vote_D"),
    ("FINAL ANSWER: correspondence", 30, "final_correspondence"),
    ("FINAL ANSWER: relation", 16, "final_relation"),
    ("Adjudication notes (why)", 30, "adj_notes"),
]
FILLIN_KEYS = {"final_correspondence", "final_relation", "adj_notes"}

INSTRUCTIONS = [
    ("What this sheet is", None),
    (None, "Four people independently labelled the same 240 items. On these "
           "43 items, the four answers didn't agree (no 3-or-4-way majority) - "
           "so these need a real decision, not an automatic tie-break."),
    ("What to do", None),
    (None, "For each row: read the old recommendation (column C) and the new "
           "edition's guideline (column D). Look at what all four annotators "
           "said (columns E-H). Decide - as a group discussion, or as "
           "whoever is adjudicating - what the ACTUAL correct answer is."),
    (None, "Then fill in columns I-K (yellow): the final correspondence "
           "(item ID, NONE, or CANNOT_DETERMINE), the final relation, and a "
           "short note on why - especially useful if the four answers were "
           "split for an interesting reason."),
    ("Important", None),
    (None, "The four annotators' original answers (E-H) are shown so you can "
           "see exactly where the disagreement is. This is different from the "
           "first round - there, everyone worked blind. Here, seeing "
           "everyone's answer is the point."),
    (None, "If you genuinely cannot resolve one after real discussion, it's "
           "fine to write CANNOT_DETERMINE as the final answer too - that's "
           "a legitimate outcome, not a failure to adjudicate."),
]


def load_report() -> dict:
    with open(BASE / "4rater_analysis_report.json", encoding="utf-8") as f:
        return json.load(f)


def load_pair_source(slug: str):
    with open(BASE / slug / "annotation_packet.csv", encoding="utf-8") as f:
        rows = {r["sample_id"]: r for r in csv.DictReader(f)}
    with open(BASE / slug / "annotation_context.json", encoding="utf-8") as f:
        ctx = json.load(f)
    return rows, ctx


def render_new_guideline(ctx_entry: dict) -> str:
    items = ctx_entry.get("new_guideline_full_text", [])
    if not items:
        return "(no items found in this guideline in the new edition)"
    blocks = [f"[{it['item_id']}]\n{it['text']}" for it in items]
    return sanitize("\n\n".join(blocks))


def vote_text(raw_entry: dict) -> str:
    corr = raw_entry.get("correspondence") or ""
    rel = raw_entry.get("relation") or ""
    return f"{corr}" + (f"  ({rel})" if rel else "")


def build_instructions_sheet(wb: Workbook):
    ws = wb.active
    ws.title = "READ ME FIRST"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 100
    r = 1
    c = ws.cell(row=r, column=1, value="Adjudication - 43 disputed items")
    c.font = Font(name=FONT_NAME, bold=True, size=16, color="C00000")
    r += 2
    for heading, body in INSTRUCTIONS:
        if heading is not None:
            c = ws.cell(row=r, column=1, value=heading)
            c.font = Font(name=FONT_NAME, bold=True, size=13, color="1F4E78")
            r += 1
        if body is not None:
            c = ws.cell(row=r, column=1, value=body)
            c.font = Font(name=FONT_NAME, size=11.5)
            c.alignment = Alignment(wrap_text=True, vertical="top")
            ws.row_dimensions[r].height = 34
            r += 1


def build_data_sheet(wb: Workbook, report: dict, raw: dict):
    ws = wb.create_sheet(title="Adjudicate these 43")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"

    for col_idx, (header, width, _key) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = WRAP_TOP
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[1].height = 30

    dv = DataValidation(type="list", formula1=f'"{RELATION_OPTIONS}"', allow_blank=True)
    dv.error = "Pick one from the list."
    dv.errorTitle = "Invalid entry"
    ws.add_data_validation(dv)

    r = 2
    for pair_title, slug in PAIR_SLUGS.items():
        sids = report["pairs"][pair_title]["needs_adjudication_sample_ids"]
        if not sids:
            continue
        src_rows, ctx = load_pair_source(slug)
        for sid in sids:
            src = src_rows[sid]
            values = {
                "pair": pair_title,
                "sample_id": sid,
                "old_text": sanitize(src["old_text"]),
                "new_guideline_text": render_new_guideline(ctx.get(sid, {})),
                "vote_A": sanitize(vote_text(raw["A"][pair_title][sid])),
                "vote_B": sanitize(vote_text(raw["B"][pair_title][sid])),
                "vote_C": sanitize(vote_text(raw["C"][pair_title][sid])),
                "vote_D": sanitize(vote_text(raw["D"][pair_title][sid])),
                "final_correspondence": "",
                "final_relation": "",
                "adj_notes": "",
            }
            for col_idx, (_h, _w, key) in enumerate(COLUMNS, start=1):
                cell = ws.cell(row=r, column=col_idx, value=values[key])
                cell.border = BORDER
                cell.alignment = WRAP_TOP
                if key in FILLIN_KEYS:
                    cell.fill = FILLIN_FILL
                elif key.startswith("vote_"):
                    cell.fill = VOTE_FILL
                    cell.font = NORMAL_FONT
                else:
                    cell.font = NORMAL_FONT
            ws.row_dimensions[r].height = 60
            dv.add(ws.cell(row=r, column=10))  # column J = final_relation
            r += 1

    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{r - 1}"
    return r - 2  # number of data rows written


if __name__ == "__main__":
    report = load_report()
    raw = {label: load_completed_xlsx(path) for label, path in ANNOTATOR_FILES.items()}

    wb = Workbook()
    build_instructions_sheet(wb)
    n = build_data_sheet(wb, report, raw)

    out_path = BASE / "Adjudication_43_items.xlsx"
    wb.save(out_path)
    print(f"wrote {out_path} ({n} rows)")
