"""
One-off script: builds the adjudication workbook for the 5 items where
Annotator G and Annotator H disagreed on the H3' 92-item packet
(2026-08-23 completion) - all 5 are the same pattern, G said NONE
(confidently deleted) while H said CANNOT_DETERMINE (genuinely unsure),
a real interpretive disagreement, not noise.

Reuses build_adjudication_sheet.py's styling/layout conventions
(instructions sheet, vote columns, yellow fill-in columns) adapted for
2 raters instead of 4.

Not part of the pipeline itself - a deliverable-formatting step over
already-collected annotation data.
"""
import csv
import json
import re
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

BASE = Path(__file__).parent
sys.path.insert(0, str(BASE.parents[3]))  # backend/ on sys.path
from app.research.cross_edition.annotation import load_completed_xlsx, _norm_answer  # noqa: E402

_ILLEGAL_XML_CHARS_RE = re.compile(
    "[\x00-\x08\x0b\x0c\x0e-\x1f\x7f-\x84\x86-\x9f\ud800-\udfff﷐-﷯￾￿]"
)


def sanitize(text) -> str:
    if not isinstance(text, str):
        return text
    return _ILLEGAL_XML_CHARS_RE.sub("", text)


PACKET_DIR = BASE / "h3prime_tennessee_2022_2024"
ANNOTATOR_G = PACKET_DIR / "Annotator_G_ANNOTATION.xlsx"
ANNOTATOR_H = PACKET_DIR / "Annotator_H_ANNOTATION.xlsx"

FONT_NAME = "Calibri"
HEADER_FILL = PatternFill("solid", fgColor="C00000")
HEADER_FONT = Font(name=FONT_NAME, bold=True, color="FFFFFF", size=11)
VOTE_FILL = PatternFill("solid", fgColor="F2F2F2")
FILLIN_FILL = PatternFill("solid", fgColor="FFF2CC")
NORMAL_FONT = Font(name=FONT_NAME, size=11)
WRAP_TOP = Alignment(wrap_text=True, vertical="top")
THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
RELATION_OPTIONS = "unchanged,reworded,substantive,merged,split,moved"

COLUMNS = [
    ("sample_id", 9, "sample_id"),
    ("OLD RECOMMENDATION", 40, "old_text"),
    ("NEW EDITION - full matching guideline", 50, "new_guideline_text"),
    ("Annotator G said", 25, "vote_G"),
    ("Annotator H said", 25, "vote_H"),
    ("FINAL ANSWER: correspondence", 30, "final_correspondence"),
    ("FINAL ANSWER: relation", 16, "final_relation"),
    ("Adjudication notes (why)", 30, "adj_notes"),
]
FILLIN_KEYS = {"final_correspondence", "final_relation", "adj_notes"}

INSTRUCTIONS = [
    ("What this sheet is", None),
    (None, "Two people independently labelled the same 92-item H3' follow-up "
           "packet. On these 5 items, their answers disagreed - both said "
           "the old recommendation had no correspondence, but one said NONE "
           "(confidently deleted) and the other said CANNOT_DETERMINE "
           "(genuinely unsure) - so these need a real decision, not an "
           "automatic tie-break."),
    ("What to do", None),
    (None, "For each row: read the old recommendation (column B) and the new "
           "edition's guideline (column C). Look at what both annotators "
           "said (columns D-E). Decide what the ACTUAL correct answer is."),
    (None, "Then fill in columns F-H (yellow): the final correspondence "
           "(item ID, NONE, or CANNOT_DETERMINE), the final relation, and a "
           "short note on why."),
    ("Important", None),
    (None, "If you genuinely cannot resolve one, CANNOT_DETERMINE as the "
           "final answer is a legitimate outcome, not a failure to "
           "adjudicate."),
]


def render_new_guideline(ctx_entry: dict) -> str:
    items = ctx_entry.get("new_guideline_full_text", [])
    if not items:
        return "(no items found in this guideline in the new edition)"
    blocks = [f"[{it['item_id']}]\n{it['text']}" for it in items]
    return sanitize("\n\n".join(blocks))


def vote_text(raw_entry: dict) -> str:
    corr = raw_entry.get("correspondence") or ""
    rel = raw_entry.get("relation") or ""
    return f"{corr}" + (f"  ({rel})" if rel else "")


def build_instructions_sheet(wb: Workbook):
    ws = wb.active
    ws.title = "READ ME FIRST"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 100
    r = 1
    c = ws.cell(row=r, column=1, value="H3' adjudication - 5 disputed items (G/H)")
    c.font = Font(name=FONT_NAME, bold=True, size=16, color="C00000")
    r += 2
    for heading, body in INSTRUCTIONS:
        if heading is not None:
            c = ws.cell(row=r, column=1, value=heading)
            c.font = Font(name=FONT_NAME, bold=True, size=13, color="1F4E78")
            r += 1
        if body is not None:
            c = ws.cell(row=r, column=1, value=body)
            c.font = Font(name=FONT_NAME, size=11.5)
            c.alignment = Alignment(wrap_text=True, vertical="top")
            ws.row_dimensions[r].height = 34
            r += 1


def build_data_sheet(wb: Workbook, disputed_sids: list[str], raw_g: dict, raw_h: dict,
                      packet: dict, ctx: dict, sheet_title: str):
    ws = wb.create_sheet(title="Adjudicate these 5")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"

    for col_idx, (header, width, _key) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = WRAP_TOP
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[1].height = 30

    dv = DataValidation(type="list", formula1=f'"{RELATION_OPTIONS}"', allow_blank=True)
    dv.error = "Pick one from the list."
    dv.errorTitle = "Invalid entry"
    ws.add_data_validation(dv)

    r = 2
    for sid in disputed_sids:
        src = packet[sid]
        values = {
            "sample_id": sid,
            "old_text": sanitize(src["old_text"]),
            "new_guideline_text": render_new_guideline(ctx.get(sid, {})),
            "vote_G": sanitize(vote_text(raw_g[sheet_title][sid])),
            "vote_H": sanitize(vote_text(raw_h[sheet_title][sid])),
            "final_correspondence": "",
            "final_relation": "",
            "adj_notes": "",
        }
        for col_idx, (_h, _w, key) in enumerate(COLUMNS, start=1):
            cell = ws.cell(row=r, column=col_idx, value=values[key])
            cell.border = BORDER
            cell.alignment = WRAP_TOP
            if key in FILLIN_KEYS:
                cell.fill = FILLIN_FILL
            elif key.startswith("vote_"):
                cell.fill = VOTE_FILL
                cell.font = NORMAL_FONT
            else:
                cell.font = NORMAL_FONT
        ws.row_dimensions[r].height = 60
        dv.add(ws.cell(row=r, column=7))  # column G = final_relation
        r += 1

    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{r - 1}"
    return r - 2


if __name__ == "__main__":
    raw_g = load_completed_xlsx(str(ANNOTATOR_G))
    raw_h = load_completed_xlsx(str(ANNOTATOR_H))
    sheet_title = list(raw_g.keys())[0]

    ga = {sid: _norm_answer(v["correspondence"]) for sid, v in raw_g[sheet_title].items()}
    ha = {sid: _norm_answer(v["correspondence"]) for sid, v in raw_h[sheet_title].items()}
    disputed = sorted(sid for sid in ga if ga[sid] != ha[sid])
    print(f"disputed items: {disputed}")

    with open(PACKET_DIR / "annotation_packet.csv", encoding="utf-8") as f:
        packet = {row["sample_id"]: row for row in csv.DictReader(f)}
    with open(PACKET_DIR / "annotation_context.json", encoding="utf-8") as f:
        ctx = json.load(f)

    wb = Workbook()
    build_instructions_sheet(wb)
    n = build_data_sheet(wb, disputed, raw_g, raw_h, packet, ctx, sheet_title)

    out_path = PACKET_DIR / "H3prime_Adjudication_5_items.xlsx"
    wb.save(out_path)
    print(f"wrote {out_path} ({n} rows)")
