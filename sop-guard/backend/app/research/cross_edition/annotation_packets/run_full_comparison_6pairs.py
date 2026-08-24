"""
One-off driver: formal re-run of H3/H4/H5 against the expanded 6-pair,
360-item confirmatory dataset - the follow-up flagged as outstanding
every time the fifth/sixth pairs were discussed (PREREGISTRATION.md's
2026-08-23 "Sample drawn..." entry and its successors): "H3/H4/H5 and
every other already-reported figure stand exactly as reported on the
original 4-pair, 240-item dataset until a formal re-run against the
expanded 6-pair, 360-item dataset is separately committed and
performed."

Extends run_full_comparison.py's exact method (same bootstrap_stat,
same stat functions, same BH correction, same index-join fix from
sample_join.py) to the two new pairs (Tennessee Sept2024->09.11.2025,
Connecticut v2024.1->v2025.1), whose ground truth comes from Annotator
I/J (2 raters, majority_vote reducing to "both agree") instead of the
original round's 4 raters.

PRELIMINARY: 6 of 120 new-pair items (3 per pair) are still pending
human adjudication (NewPairs_Adjudication_6_items.xlsx, generated,
awaiting completion) - excluded from ground truth here, not guessed
at, exactly the treatment already established for H3' and for the
new pairs' own section 6 metrics (run_new_pairs_metrics.py). This
script's headline pooled numbers should be re-run once those resolve.

Reports BOTH the original 4-pair pooled result (recomputed here,
unchanged in method) and the new 6-pair pooled result side by side,
per section 10's standing preference for showing rather than asserting
a difference's size, plus a 2-pair-only breakout for the new pairs on
their own so the marginal contribution is visible.

Does not modify item_align.py, item_parser.py, corpus_probe.py, or
edition_align.py.

Run:
    cd sop-guard/backend
    python -m app.research.cross_edition.annotation_packets.run_full_comparison_6pairs
"""
from __future__ import annotations

import csv
import json
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[4]))  # backend/ on sys.path

from app.research.cross_edition.annotation import (  # noqa: E402
    load_completed_xlsx, _norm_answer, majority_vote,
)
from app.research.cross_edition.baseline_b2 import align_items_b2  # noqa: E402
from app.research.cross_edition.baseline_b1_b3_b4 import (  # noqa: E402
    align_items_b1, align_items_b3, align_items_b4,
)
from app.research.cross_edition.annotation_packets.sample_join import (  # noqa: E402
    build_index_join, verify_sample_identity, join_baseline,
)
from app.research.cross_edition.annotation_packets.run_h3_test import (  # noqa: E402
    PAIRS as ORIGINAL_PAIRS, build_ground_truth as build_original_ground_truth,
)
import app.research.cross_edition.annotation_packets.run_full_comparison as fc  # noqa: E402

BASE = Path(__file__).parent
SP = r"C:\Users\Faisal\Desktop\Hospital SOP's Research\corpus\protocols"

NEW_PAIRS = {
    "Tennessee Sept2024→09.11.2025 (": (
        "tennessee_sept2024_20250911", f"{SP}\\tn_sept2024.pdf", f"{SP}\\tn_20250911.pdf"),
    "Connecticut v2024.1→v2025.1 (ne": (
        "connecticut_v20241_v20251", f"{SP}\\ct_v20241.pdf", f"{SP}\\ct_v20251.pdf"),
}

ANNOTATOR_I = BASE / "Annotator_I_ANNOTATION.xlsx"
ANNOTATOR_J = BASE / "Annotator_J_ANNOTATION.xlsx"
ADJUDICATION_FILE = BASE / "NewPairs_Adjudication_6_items_COMPLETED.xlsx"


def _load_new_pairs_adjudication() -> dict[str, dict[str, str]]:
    out: dict[str, dict[str, str]] = {p: {} for p in NEW_PAIRS}
    if not ADJUDICATION_FILE.exists():
        return out
    from openpyxl import load_workbook
    wb = load_workbook(ADJUDICATION_FILE, data_only=True)
    ws = wb["Adjudicate these 6"]
    for r in range(2, ws.max_row + 1):
        pair = ws.cell(row=r, column=1).value
        sid = ws.cell(row=r, column=2).value
        final_corr = ws.cell(row=r, column=7).value
        if pair in out and sid and final_corr:
            out[pair][sid] = _norm_answer(final_corr)
    return out


def build_new_pairs_ground_truth() -> tuple[dict[str, dict[str, str]], dict[str, list[str]]]:
    raw_i = load_completed_xlsx(str(ANNOTATOR_I))
    raw_j = load_completed_xlsx(str(ANNOTATOR_J))
    adjudicated = _load_new_pairs_adjudication()

    ground_truth: dict[str, dict[str, str]] = {}
    pending: dict[str, list[str]] = {}
    for pair in NEW_PAIRS:
        norm_i = {sid: _norm_answer(v["correspondence"]) for sid, v in raw_i[pair].items()}
        norm_j = {sid: _norm_answer(v["correspondence"]) for sid, v in raw_j[pair].items()}
        mv = majority_vote([norm_i, norm_j])
        gt, still_pending = {}, []
        for sid, result in mv.items():
            if result["answer"] is not None:
                gt[sid] = result["answer"]
            elif sid in adjudicated[pair]:
                gt[sid] = adjudicated[pair][sid]
            else:
                still_pending.append(sid)
        ground_truth[pair] = gt
        pending[pair] = sorted(still_pending)
    return ground_truth, pending


def build_records_for(pair_slugs: dict, ground_truth: dict, start_idx: int = 0) -> list[dict]:
    """Mirrors run_full_comparison.build_records() exactly (same index-join,
    same fail-loud verify, same 4-baseline computation), generalized to take
    an arbitrary {pair_title: (slug, old_pdf, new_pdf)} map and ground truth
    so it works for both the original and new pairs without duplicating the
    join logic."""
    records: list[dict] = []
    for offset, (pair, (slug, old_pdf, new_pdf)) in enumerate(pair_slugs.items()):
        pair_idx = start_idx + offset
        packet_csv = BASE / slug / "annotation_packet.csv"
        with open(packet_csv, encoding="utf-8") as f:
            method_rows = {r["sample_id"]: r for r in csv.DictReader(f)}

        id_to_index, _ = build_index_join(old_pdf, new_pdf)
        verify_sample_identity(packet_csv, id_to_index)

        b1 = align_items_b1(old_pdf, new_pdf)
        b2 = align_items_b2(old_pdf, new_pdf)
        b3 = align_items_b3(old_pdf, new_pdf)
        b4 = align_items_b4(old_pdf, new_pdf)

        gt = ground_truth[pair]
        for sid, row in method_rows.items():
            if sid not in gt:
                continue
            truth = gt[sid]
            if truth == "cannot_determine":
                continue
            idx = id_to_index[row["old_item_id"]]

            method_pred = _norm_answer(row["method_predicted_item_id"])
            preds = {
                "method": method_pred,
                "b1": _norm_answer(join_baseline(b1, idx, "b1_predicted_item_id")),
                "b2": _norm_answer(join_baseline(b2, idx, "b2_predicted_item_id")),
                "b3": _norm_answer(join_baseline(b3, idx, "b3_predicted_item_id")),
                "b4": _norm_answer(join_baseline(b4, idx, "b4_predicted_item_id")),
            }
            weight = float(row.get("sample_weight", 1.0) or 1.0)

            records.append({
                "pair": pair, "pair_idx": pair_idx, "sample_id": sid,
                "tier": row["tier"], "truth": truth, "weight": weight,
                "preds": preds,
                "correct": {k: (1 if v == truth else 0) for k, v in preds.items()},
            })
    return records


def run_hypothesis_tests(records: list[dict], label: str) -> dict:
    """Mirrors run_full_comparison.main()'s body exactly, factored out so it
    can run once on the original 4 pairs, once on the new 2, and once on
    the pooled 6 - all via literally the same stat functions and BH code."""
    result: dict = {"label": label, "n_items": len(records)}

    h5 = {}
    for weighted, wlabel in ((False, "raw"), (True, "weighted")):
        h5[wlabel] = fc.bootstrap_stat(records, fc.make_false_corr_diff_fn("method", "b1", weighted))
    result["H5_false_correspondence_method_minus_B1"] = h5
    result["H5_confirmed"] = {
        "item_raw": h5["raw"]["item"]["ci95_high"] is not None and h5["raw"]["item"]["ci95_high"] < 0.05,
        "pair_raw": h5["raw"]["pair"]["ci95_high"] is not None and h5["raw"]["pair"]["ci95_high"] < 0.05,
    }

    h3 = {}
    for weighted, wlabel in ((False, "raw"), (True, "weighted")):
        h3[wlabel] = fc.bootstrap_stat(records, fc.make_accuracy_diff_fn("method", "b2", weighted))
    result["H3_method_minus_B2_accuracy"] = h3
    result["H3_confirmed"] = {
        "item_raw": h3["raw"]["item"]["point_estimate"] is not None and h3["raw"]["item"]["point_estimate"] > 0
                    and h3["raw"]["item"]["ci95_low"] is not None and h3["raw"]["item"]["ci95_low"] > 0,
        "pair_raw": h3["raw"]["pair"]["point_estimate"] is not None and h3["raw"]["pair"]["point_estimate"] > 0
                    and h3["raw"]["pair"]["ci95_low"] is not None and h3["raw"]["pair"]["ci95_low"] > 0,
    }

    t3_point = fc.t3_precision_fn(records)
    result["H4_T3_precision"] = {"point_estimate": round(t3_point, 4) if t3_point is not None else None}
    result["H4_confirmed"] = t3_point is not None and t3_point >= 0.80

    desc = {}
    for other in ("b3", "b4"):
        d = {}
        for weighted, wlabel in ((False, "raw"), (True, "weighted")):
            d[wlabel] = fc.bootstrap_stat(records, fc.make_accuracy_diff_fn("method", other, weighted))
        desc[f"method_minus_{other}_accuracy"] = d
    result["descriptive"] = desc

    pvals = {
        "H3": fc.bootstrap_p_h3(records),
        "H4": fc.bootstrap_p_h4(records),
        "H5": fc.bootstrap_p_h5(records),
    }
    result["benjamini_hochberg"] = fc.benjamini_hochberg(pvals)

    print(f"=== {label} (n={len(records)}) ===")
    print(f"  H3 (method-B2 accuracy): item-raw={h3['raw']['item']}  pair-raw={h3['raw']['pair']}  "
          f"confirmed={result['H3_confirmed']}")
    print(f"  H4 (T3 precision): {t3_point}  confirmed={result['H4_confirmed']}")
    print(f"  H5 (method-B1 false-corr): item-raw={h5['raw']['item']}  pair-raw={h5['raw']['pair']}  "
          f"confirmed={result['H5_confirmed']}")
    print(f"  BH-adjusted: {result['benjamini_hochberg']}")
    print()
    return result


def main():
    original_gt = build_original_ground_truth()
    new_gt, pending = build_new_pairs_ground_truth()

    total_pending = sum(len(v) for v in pending.values())
    print(f"new-pair items pending adjudication (excluded below): {total_pending} - {pending}")

    original_records = build_records_for(ORIGINAL_PAIRS, original_gt, start_idx=0)
    new_records = build_records_for(NEW_PAIRS, new_gt, start_idx=len(ORIGINAL_PAIRS))
    print(f"original 4-pair usable items: {len(original_records)}  "
          f"new 2-pair usable items: {len(new_records)}")

    report = {
        "_preliminary_note": (
            f"{total_pending} of 120 new-pair items are still pending human "
            "adjudication (NewPairs_Adjudication_6_items.xlsx) and are "
            "excluded from every number below, not estimated. Re-run once "
            "resolved."
        ),
        "pending_adjudication": pending,
        "original_4pairs": run_hypothesis_tests(original_records, "ORIGINAL 4 PAIRS (240 sampled, recomputed here)"),
        "new_2pairs": run_hypothesis_tests(new_records, "NEW 2 PAIRS ONLY (120 sampled)"),
        "pooled_6pairs": run_hypothesis_tests(original_records + new_records, "POOLED 6 PAIRS (360 sampled)"),
    }

    out_path = BASE / "full_comparison_6pairs_report.json"
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(report, f, indent=2, ensure_ascii=False)
    print(f"wrote {out_path}")


if __name__ == "__main__":
    main()
