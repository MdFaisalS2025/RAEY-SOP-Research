"""
One-off driver: assembles ground truth for the seventh/eighth confirmatory
pairs (PREREGISTRATION.md's 2026-08-24 "Seventh and eighth confirmatory
pairs" entry - Massachusetts v2025.1->v2026.1, v2026.1->v2026.2) from the
completed Annotator K / Annotator L workbooks, and computes section 6's
metrics against the method's predictions. Mirrors
run_new_pairs_metrics.py's structure exactly (same 2-rater
majority_vote/pending-adjudication treatment, same independence check).

Not part of the pipeline itself - a report generator over already-
collected annotation data. Does not modify item_align.py,
item_parser.py, corpus_probe.py, or edition_align.py. (item_parser_ma.py
is not touched here either - the method_predicted_item_id column is
already baked into each pair's annotation_packet.csv from the sample
draw, so this script only needs the frozen `annotation` module's
generic scoring helpers, not the MA-specific parser itself.)

Run:
    cd sop-guard/backend
    python -m app.research.cross_edition.annotation_packets.run_ma_pairs_metrics
"""
from __future__ import annotations

import csv
import hashlib
import json
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[4]))  # backend/ on sys.path

from app.research.cross_edition.annotation import (  # noqa: E402
    load_completed_xlsx, _norm_answer, majority_vote, compute_section6_metrics,
)
from app.research.cross_edition.annotation_packets.run_4rater_analysis import cohens_kappa  # noqa: E402

BASE = Path(__file__).parent

ANNOTATOR_K = BASE / "Annotator_K_ANNOTATION.xlsx"
ANNOTATOR_L = BASE / "Annotator_L_ANNOTATION.xlsx"
ADJUDICATION_FILE = BASE / "MAPairs_Adjudication_2_items_COMPLETED.xlsx"  # optional, generated on demand

PAIR_SLUGS = {
    "Massachusetts v2025.1→v2026.1 (": "massachusetts_v20251_v20261",
    "Massachusetts v2026.1→v2026.2 (": "massachusetts_v20261_v20262",
}


def verify_independence(path_k: Path, path_l: Path) -> dict:
    """Mirrors run_new_pairs_metrics.py's verify_independence exactly."""
    hash_k = hashlib.md5(path_k.read_bytes()).hexdigest()
    hash_l = hashlib.md5(path_l.read_bytes()).hexdigest()
    identical_bytes = hash_k == hash_l

    raw_k = load_completed_xlsx(str(path_k))
    raw_l = load_completed_xlsx(str(path_l))
    per_pair = {}
    all_identical = True
    for sheet in PAIR_SLUGS:
        ans_k = {sid: _norm_answer(v["correspondence"]) for sid, v in raw_k.get(sheet, {}).items()}
        ans_l = {sid: _norm_answer(v["correspondence"]) for sid, v in raw_l.get(sheet, {}).items()}
        identical = ans_k == ans_l and len(ans_k) > 0
        per_pair[sheet] = {"n_k": len(ans_k), "n_l": len(ans_l), "identical_answers": identical}
        all_identical = all_identical and identical

    report = {
        "hash_k": hash_k, "hash_l": hash_l, "identical_bytes": identical_bytes,
        "per_pair": per_pair, "identical_answers_all_pairs": all_identical,
    }
    print(f"Independence check: {json.dumps(report, indent=2)}")
    if identical_bytes:
        raise RuntimeError("Annotator K/L files are byte-identical - collection "
                            "failure, not agreement. Do not proceed.")
    if all_identical:
        raise RuntimeError("Annotator K/L answers are identical on every pair despite "
                            "different file bytes - the same pattern that caught the "
                            "original E/F and A/B duplications. Treat as unverified "
                            "until re-collected.")
    return report


def _load_ma_pairs_adjudication(path: Path) -> dict[str, dict[str, str]]:
    """Mirrors run_new_pairs_metrics.py's _load_new_pairs_adjudication
    exactly - reads a from-scratch adjudication workbook's own column
    layout directly, not via load_completed_xlsx."""
    from openpyxl import load_workbook
    wb = load_workbook(path, data_only=True)
    ws = wb.worksheets[-1]
    out: dict[str, dict[str, str]] = {p: {} for p in PAIR_SLUGS}
    header = [c.value for c in ws[1]]
    final_col = None
    for idx, h in enumerate(header, start=1):
        if h and "final" in str(h).lower() and "correspond" in str(h).lower():
            final_col = idx
    if final_col is None:
        final_col = ws.max_column
    for r in range(2, ws.max_row + 1):
        pair = ws.cell(row=r, column=1).value
        sid = ws.cell(row=r, column=2).value
        final_corr = ws.cell(row=r, column=final_col).value
        if pair in out and sid and final_corr:
            out[pair][sid] = _norm_answer(final_corr)
    return out


def build_ground_truth() -> tuple[dict[str, dict[str, str]], dict[str, list[str]]]:
    raw_k = load_completed_xlsx(str(ANNOTATOR_K))
    raw_l = load_completed_xlsx(str(ANNOTATOR_L))

    adjudicated = {}
    if ADJUDICATION_FILE.exists():
        adjudicated = _load_ma_pairs_adjudication(ADJUDICATION_FILE)

    ground_truth: dict[str, dict[str, str]] = {}
    pending: dict[str, list[str]] = {}
    for sheet in PAIR_SLUGS:
        norm_k = {sid: _norm_answer(v["correspondence"]) for sid, v in raw_k[sheet].items()}
        norm_l = {sid: _norm_answer(v["correspondence"]) for sid, v in raw_l[sheet].items()}
        mv = majority_vote([norm_k, norm_l])
        gt, still_pending = {}, []
        for sid, result in mv.items():
            if result["answer"] is not None:
                gt[sid] = result["answer"]
            else:
                adj = adjudicated.get(sheet, {}).get(sid)
                if adj is not None:
                    gt[sid] = adj
                else:
                    still_pending.append(sid)
        ground_truth[sheet] = gt
        pending[sheet] = sorted(still_pending)
    return ground_truth, pending


def load_method_rows(slug: str) -> list[dict]:
    with open(BASE / slug / "annotation_packet.csv", encoding="utf-8") as f:
        return list(csv.DictReader(f))


def main():
    if not ANNOTATOR_K.exists() or not ANNOTATOR_L.exists():
        print(f"Waiting for both files to exist at:\n  {ANNOTATOR_K}\n  {ANNOTATOR_L}")
        return

    verify_independence(ANNOTATOR_K, ANNOTATOR_L)

    raw_k = load_completed_xlsx(str(ANNOTATOR_K))
    raw_l = load_completed_xlsx(str(ANNOTATOR_L))

    ground_truth, pending = build_ground_truth()

    report: dict = {"pairs": {}, "pooled": {}, "kappa": {}, "pending_adjudication": pending}
    pooled_gt: dict[str, str] = {}
    pooled_rows: list[dict] = []
    pooled_k: dict[str, str] = {}
    pooled_l: dict[str, str] = {}

    for sheet, slug in PAIR_SLUGS.items():
        norm_k = {sid: _norm_answer(v["correspondence"]) for sid, v in raw_k[sheet].items()}
        norm_l = {sid: _norm_answer(v["correspondence"]) for sid, v in raw_l[sheet].items()}
        ck = cohens_kappa(norm_k, norm_l)
        report["kappa"][sheet] = ck
        print(f"=== {sheet} ===")
        print(f"  Cohen's kappa (K/L): {ck['cohens_kappa']}  "
              f"(observed agreement {ck['observed_agreement']}, "
              f"{len(ck['disagreements'])}/{ck['n']} disagreements)")
        if pending[sheet]:
            print(f"  PENDING ADJUDICATION (excluded from ground truth): "
                  f"{len(pending[sheet])} items - {pending[sheet]}")

        rows = load_method_rows(slug)
        metrics = compute_section6_metrics(ground_truth[sheet], rows)
        report["pairs"][sheet] = metrics
        print(json.dumps(metrics, indent=2))
        print()

        for sid, ans in ground_truth[sheet].items():
            key = f"{sheet}::{sid}"
            pooled_gt[key] = ans
        for r in rows:
            r2 = dict(r)
            r2["sample_id"] = f"{sheet}::{r['sample_id']}"
            pooled_rows.append(r2)
        for sid, v in norm_k.items():
            pooled_k[f"{sheet}::{sid}"] = v
        for sid, v in norm_l.items():
            pooled_l[f"{sheet}::{sid}"] = v

    pooled_ck = cohens_kappa(pooled_k, pooled_l)
    report["kappa"]["pooled"] = pooled_ck
    pooled_metrics = compute_section6_metrics(pooled_gt, pooled_rows)
    report["pooled"] = pooled_metrics

    print("=== POOLED (both MA pairs, 120 items) ===")
    print(f"  Cohen's kappa (K/L, pooled): {pooled_ck['cohens_kappa']}  "
          f"(observed agreement {pooled_ck['observed_agreement']}, "
          f"{len(pooled_ck['disagreements'])}/{pooled_ck['n']} disagreements)")
    print(json.dumps(pooled_metrics, indent=2))

    out_path = BASE / "ma_pairs_final_metrics.json"
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(report, f, indent=2, ensure_ascii=False)
    print(f"\nwrote {out_path}")


if __name__ == "__main__":
    main()
