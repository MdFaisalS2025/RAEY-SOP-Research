"""
Annotation instrument: stratified sampling, packet generation, kappa.

STATUS AS OF WRITING: real confirmatory data now exists. PREREGISTRATION.md
section 11's 2026-08-17 entries (the Connecticut-reset entry and the two
that follow it) record that four edition pairs meet section 3.2's
minimum-viable confirmatory test set (>=4 pairs, >=3 publishers), verified
against git history and each publisher's own official records, and
classified from each document's own front matter (not inferred) as minor
revisions:

    - Tennessee 2017 -> 2018
    - Pennsylvania 2021 -> 2023v1-2
    - Connecticut v2022.1 -> v2023.1
    - Connecticut v2023.1 -> v2024.1

Sampling and packets generated from these four pairs ARE confirmatory,
not a dry run, and may be reported as such once independently annotated
per section 5.3.

One live limitation: all four pairs are MINOR revisions. H1 and H2 (section
7) require a MAJOR revision pair, and a genuine, criteria-gated search
(PREREGISTRATION.md section 11, 2026-08-17 stopping-rule entries) found
none available and was stopped deliberately rather than left open-ended.
H1 and H2 are therefore untestable with the current dataset - not
disconfirmed, simply without the required stratum. H3 and H4 do not
require a major/minor split and are testable on the four pairs above.

This module may still be exercised against dev data (NASEMSO, or the
NY/Maine editions that showed clean parsing but poor alignment quality -
see FEASIBILITY.md section 42) as a mechanics check - every function that
touches real documents says so in its output - but nothing produced against
dev data may be treated as, or reported as, confirmatory annotation.

WHAT THIS MODULE CANNOT DO: section 5.3 requires two annotators to label
independently, having "not seen any method output for the item they are
labelling." Whoever built and inspected this pipeline's output extensively
throughout retrieval and testing cannot serve as one of those two
annotators without violating that independence requirement. This module
prepares the sample and the packets; it does not and cannot perform the
actual labelling.

WHAT THIS IMPLEMENTS, AND WHERE IT COMES FROM
----------------------------------------------
PREREGISTRATION.md section 5.1 (sampling): a stratified random sample of 60
old items per pair, allocated across the tiers item_align.py assigns - 12
per tier for T1-T5, any T6 shortfall or surplus redistributed
proportionally - because ground truth is annotated, not computed (section
3.2's revision-date shortcut was tested and disconfirmed: see
FEASIBILITY.md section 3.2 / PREREGISTRATION.md's original 2026-08-16
entry).

section 5.2 (task): for each sampled old item, an annotator sees the item in
context plus the WHOLE corresponding new-edition guideline, and records
either the corresponding new item, NONE (deleted), or CANNOT DETERMINE, plus
a relation label (unchanged / reworded only / substantively changed /
merged / split / moved).

section 5.3 (agreement): two annotators label independently; Cohen's kappa
on the pre-adjudication correspondence judgment is the reported reliability
statistic.

Run:
    cd sop-guard/backend
    python -m app.research.cross_edition.annotation sample old.pdf new.pdf out_dir
    python -m app.research.cross_edition.annotation kappa packet_a.csv packet_b.csv
"""

from __future__ import annotations

import csv
import json
import random
import sys
from collections import defaultdict
from pathlib import Path

from app.research.cross_edition.item_align import align_items, _norm
from app.research.cross_edition.item_parser import Item

# section 5.1's per-tier target and overall target.
#
# CORRECTED 2026-08-17. The pre-registration originally read "12 per tier
# T1-T5, T6 shortfall redistributed" - internally inconsistent, since 12x5
# tiers already equals the stated 60-item total with T6 excluded from the
# flat allocation entirely. A first implementation followed that literally
# (12 from EACH of six tiers, T6 included) and drew 72, not 60 - caught by
# checking total_drawn against the stated target rather than trusting the
# number. T6 cannot be excluded in practice: PREREGISTRATION.md section 6
# names deletion recall/precision as a primary metric, which needs T6
# samples to exist at all. Fixed to 10 items x 6 tiers = 60, T6 included in
# the flat allocation on equal footing with T1-T5. See PREREGISTRATION.md
# section 5.1 and its section 11 deviation entry for the corrected text.
_PER_TIER_TARGET = 10
_TOTAL_TARGET = 60
_TIERS = ["T1_id_exact", "T2_id_text_changed", "T3_renumbered",
          "T4_reworded", "T5_moved", "T6_unmatched_old"]


def stratified_sample(
    old_pdf: str, new_pdf: str, seed: int = 20261017,
) -> dict:
    """Draw the section 5.1 sample from one edition pair.

    Returns a dict with the sample records, the population size per tier
    (needed to reweight any statistic computed on the sample back to the
    population, per section 5.1's "all reported rates are reweighted to the
    population" requirement), and the per-tier sampling weight.

    T6 (unmatched) has no `new_item` from the method's own assignment -
    the whole point of annotating it is to determine what the method could
    not. Those rows are still drawn and packeted; the annotator's job for
    them is exactly to say NONE (real deletion) or supply the missed
    correspondence.
    """
    result = align_items(old_pdf, new_pdf)
    all_results = result["_all_results"]

    by_tier: dict[str, list[dict]] = defaultdict(list)
    for r in all_results:
        by_tier[r["tier"]].append(r)

    rng = random.Random(seed)
    sample: list[dict] = []
    weights: dict[str, float] = {}

    # First pass: draw up to _PER_TIER_TARGET from each tier that has enough.
    shortfall = 0
    drawn_per_tier: dict[str, int] = {}
    for tier in _TIERS:
        pool = by_tier.get(tier, [])
        k = min(_PER_TIER_TARGET, len(pool))
        drawn_per_tier[tier] = k
        shortfall += _PER_TIER_TARGET - k

    # Redistribute the shortfall proportionally across tiers with spare
    # population, per section 5.1. Proportional to each tier's REMAINING
    # (undrawn) population, so a huge tier (T1 is often 40-90% of all
    # items) absorbs most of the redistribution rather than a small one
    # being oversampled into implausibility.
    if shortfall > 0:
        remaining = {t: len(by_tier.get(t, [])) - drawn_per_tier[t] for t in _TIERS}
        total_remaining = sum(max(0, v) for v in remaining.values())
        if total_remaining > 0:
            extra_allocated = 0
            for tier in _TIERS:
                if remaining[tier] <= 0:
                    continue
                extra = int(shortfall * remaining[tier] / total_remaining)
                extra = min(extra, remaining[tier])
                drawn_per_tier[tier] += extra
                extra_allocated += extra
            # Any leftover from integer rounding goes to the largest
            # remaining pool.
            leftover = shortfall - extra_allocated
            if leftover > 0:
                biggest = max(_TIERS, key=lambda t: remaining.get(t, 0) -
                              (drawn_per_tier[t] - min(_PER_TIER_TARGET, len(by_tier.get(t, [])))))
                drawn_per_tier[biggest] = min(
                    drawn_per_tier[biggest] + leftover, len(by_tier.get(biggest, []))
                )

    for tier in _TIERS:
        pool = by_tier.get(tier, [])
        k = drawn_per_tier[tier]
        if k == 0:
            weights[tier] = 0.0
            continue
        chosen = rng.sample(pool, k)
        weights[tier] = len(pool) / k  # inverse sampling fraction
        for r in chosen:
            sample.append({**r, "sample_weight": round(len(pool) / k, 4)})

    return {
        "old_pdf": old_pdf, "new_pdf": new_pdf, "seed": seed,
        "population_by_tier": {t: len(by_tier.get(t, [])) for t in _TIERS},
        "drawn_by_tier": drawn_per_tier,
        "sample_weight_by_tier": weights,
        "sample": sample,
        "total_drawn": len(sample),
        "total_target": _TOTAL_TARGET,
    }


def write_annotation_packet(sample_result: dict, out_dir: str) -> tuple[str, str]:
    """Write a human-fillable CSV plus a JSON context file with the full
    corresponding new-edition guideline for each sampled item (section
    5.2 requires the annotator see "the whole corresponding new-edition
    guideline", which does not fit in a CSV cell).

    Returns (csv_path, context_json_path).
    """
    out = Path(out_dir)
    out.mkdir(parents=True, exist_ok=True)

    from app.research.cross_edition.item_parser import parse
    new_ed = parse(sample_result["new_pdf"])
    by_guideline: dict[str, list[Item]] = defaultdict(list)
    for it in new_ed.items:
        by_guideline[_norm(it.guideline)].append(it)

    rows = []
    context = {}
    for k, rec in enumerate(sample_result["sample"]):
        sid = f"S{k+1:03d}"
        old: Item = rec["old_item"]
        new: Item | None = rec["new_item"]
        rows.append({
            "sample_id": sid,
            "tier": rec["tier"],
            "sample_weight": rec["sample_weight"],
            "old_item_id": old.item_id,
            "old_guideline": old.guideline,
            "old_section": old.section,
            "old_marker_path": old.marker_path,
            "old_text": old.text,
            "method_similarity": rec["similarity"],
            "method_predicted_item_id": new.item_id if new else "NONE",
            "method_predicted_text": new.text if new else "",
            # --- annotator fills in everything below ---
            "annotator_correspondence": "",  # new_item_id, NONE, or CANNOT_DETERMINE
            "annotator_relation": "",  # unchanged/reworded/substantive/merged/split/moved
            "annotator_notes": "",
        })
        # Full corresponding new-edition guideline, so the annotator is not
        # dependent on having the source PDF open - section 5.2's
        # requirement, made concrete.
        gkey = _norm(old.guideline)
        guideline_items = sorted(
            by_guideline.get(gkey, []), key=lambda i: i.char_start
        )
        context[sid] = {
            "old_item_full": {"id": old.item_id, "text": old.text,
                              "guideline": old.guideline, "section": old.section},
            "new_guideline_full_text": [
                {"item_id": i.item_id, "marker_path": i.marker_path, "text": i.text}
                for i in guideline_items
            ],
        }

    csv_path = out / "annotation_packet.csv"
    with open(csv_path, "w", encoding="utf-8", newline="") as f:
        w = csv.DictWriter(f, fieldnames=list(rows[0].keys()) if rows else [])
        w.writeheader()
        w.writerows(rows)

    ctx_path = out / "annotation_context.json"
    with open(ctx_path, "w", encoding="utf-8") as f:
        json.dump(context, f, indent=2)

    readme_path = out / "README.md"
    with open(readme_path, "w", encoding="utf-8") as f:
        f.write(_README_TEMPLATE.format(
            old=sample_result["old_pdf"], new=sample_result["new_pdf"],
            n=len(rows), seed=sample_result["seed"],
        ))

    return str(csv_path), str(ctx_path)


_README_TEMPLATE = """# Annotation packet

Source: {old} -> {new}
Sample size: {n} items (seed {seed}, see PREREGISTRATION.md section 5.1)

## Task (PREREGISTRATION.md section 5.2)

For each row in annotation_packet.csv, look up the sample_id in
annotation_context.json to see the full old item and the WHOLE
corresponding new-edition guideline. Decide:

1. **annotator_correspondence**: the new item_id that is the same
   recommendation as the old item, or `NONE` if it was genuinely deleted,
   or `CANNOT_DETERMINE` if you cannot tell from the documents alone.
2. **annotator_relation**: one of `unchanged`, `reworded`, `substantive`,
   `merged`, `split`, `moved`.
3. **annotator_notes**: anything worth recording, especially for
   CANNOT_DETERMINE.

Do NOT look at method_predicted_item_id before deciding - it is the
method's own guess, and seeing it first defeats the point of an independent
judgement. Cover it if annotating on paper; if annotating in a spreadsheet,
hide that column until after your first pass.

Two annotators complete this independently. Do not compare notes until
both are done - PREREGISTRATION.md section 5.3 reports agreement on the
UNINFLUENCED judgements.
"""


def compute_kappa(csv_a: str, csv_b: str) -> dict:
    """Cohen's kappa on the correspondence judgement, per section 5.3.
    No external stats dependency - the formula is direct enough to implement
    against, and this avoids adding a new package for one metric."""
    def load(path: str) -> dict[str, str]:
        with open(path, encoding="utf-8") as f:
            return {row["sample_id"]: row["annotator_correspondence"].strip()
                    for row in csv.DictReader(f)}

    a, b = load(csv_a), load(csv_b)
    common = sorted(set(a) & set(b))
    if not common:
        return {"error": "no shared sample_ids between the two packets"}

    pairs = [(a[k], b[k]) for k in common]
    n = len(pairs)
    agree = sum(1 for x, y in pairs if x == y)
    po = agree / n

    cats = sorted(set(x for x, _ in pairs) | set(y for _, y in pairs))
    pa = {c: sum(1 for x, _ in pairs if x == c) / n for c in cats}
    pb = {c: sum(1 for _, y in pairs if y == c) / n for c in cats}
    pe = sum(pa[c] * pb[c] for c in cats)

    kappa = (po - pe) / (1 - pe) if pe < 1 else float("nan")
    return {
        "n": n, "observed_agreement": round(po, 4),
        "expected_agreement": round(pe, 4), "cohens_kappa": round(kappa, 4),
        "disagreements": [
            {"sample_id": k, "a": a[k], "b": b[k]}
            for k in common if a[k] != b[k]
        ],
    }


def _norm_answer(v: str | None) -> str:
    """Normalise an annotator's typed correspondence answer for comparison
    across raters - case/whitespace differences from retyping an item_id are
    not meaningful disagreement. Section 5.2's fixed vocabulary (NONE,
    CANNOT_DETERMINE) is uppercased on comparison regardless of how it was
    typed; a copied item_id is just trimmed and case-folded."""
    if v is None:
        return ""
    v = str(v).strip()
    if v.upper() in ("NONE", "CANNOT_DETERMINE", "CANNOT DETERMINE"):
        return "CANNOT_DETERMINE" if "CANNOT" in v.upper() else "NONE"
    return v.strip().lower()


def load_completed_xlsx(path: str) -> dict[str, dict[str, dict]]:
    """Load one annotator's completed workbook (built by
    build_annotator_workbooks.py). Returns {sheet_title: {sample_id:
    {"correspondence": ..., "relation": ...}}}, both raw (as typed) and
    available for normalisation by the caller."""
    from openpyxl import load_workbook

    wb = load_workbook(path, data_only=True)
    out: dict[str, dict[str, dict]] = {}
    for sheet in wb.sheetnames:
        if sheet == "READ ME FIRST":
            continue
        ws = wb[sheet]
        pair_data = {}
        for r in range(2, ws.max_row + 1):
            sid = ws.cell(row=r, column=1).value
            if not sid:
                continue
            pair_data[sid] = {
                "correspondence": ws.cell(row=r, column=7).value,
                "relation": ws.cell(row=r, column=8).value,
                "notes": ws.cell(row=r, column=9).value,
            }
        out[sheet] = pair_data
    return out


def fleiss_kappa_correspondence(rater_answers: list[dict[str, str]]) -> dict:
    """Fleiss' kappa on the correspondence judgement across N>=2 raters -
    the correct multi-rater generalisation of Cohen's kappa (Cohen's is only
    defined for exactly two raters). rater_answers: one {sample_id: answer}
    dict per rater, already normalised by the caller (see _norm_answer).

    Standard Fleiss' kappa formula (Fleiss 1971): for each item, count votes
    per category; P_i = agreement rate for that item; P_bar = mean of P_i;
    P_e = sum of squared category proportions across the whole pool;
    kappa = (P_bar - P_e) / (1 - P_e).
    """
    if len(rater_answers) < 2:
        return {"error": "need at least 2 raters"}

    common = sorted(set.intersection(*(set(r) for r in rater_answers)))
    if not common:
        return {"error": "no shared sample_ids across all raters"}

    n_raters = len(rater_answers)
    cats = sorted({r[sid] for r in rater_answers for sid in common})
    cat_index = {c: i for i, c in enumerate(cats)}

    # n_ij: for item i, how many raters chose category j
    counts = []
    for sid in common:
        row = [0] * len(cats)
        for r in rater_answers:
            row[cat_index[r[sid]]] += 1
        counts.append(row)

    N = len(common)
    P_i = [
        (sum(c * c for c in row) - n_raters) / (n_raters * (n_raters - 1))
        for row in counts
    ]
    P_bar = sum(P_i) / N
    p_j = [sum(row[j] for row in counts) / (N * n_raters) for j in range(len(cats))]
    P_e = sum(p * p for p in p_j)

    kappa = (P_bar - P_e) / (1 - P_e) if P_e < 1 else float("nan")
    return {
        "n_items": N, "n_raters": n_raters, "categories": cats,
        "mean_observed_agreement": round(P_bar, 4),
        "expected_agreement": round(P_e, 4),
        "fleiss_kappa": round(kappa, 4),
    }


def majority_vote(rater_answers: list[dict[str, str]]) -> dict[str, dict]:
    """Per-item majority vote across N raters (already normalised answers).
    Returns {sample_id: {"answer": winning category or None, "votes": {...},
    "unanimous": bool, "needs_adjudication": bool}}. An item needs
    adjudication when no category has a strict majority (> half the raters) -
    a 2-2 split among 4, or complete four-way disagreement - per section
    5.3's adjudication step, not resolved automatically here."""
    common = sorted(set.intersection(*(set(r) for r in rater_answers)))
    n_raters = len(rater_answers)
    result = {}
    for sid in common:
        votes: dict[str, int] = {}
        for r in rater_answers:
            votes[r[sid]] = votes.get(r[sid], 0) + 1
        winner, top = max(votes.items(), key=lambda kv: kv[1])
        result[sid] = {
            "answer": winner if top > n_raters / 2 else None,
            "votes": votes,
            "unanimous": len(votes) == 1,
            "needs_adjudication": top <= n_raters / 2,
        }
    return result


def main(argv: list[str]) -> int:
    if len(argv) < 2:
        print(__doc__)
        return 2
    cmd = argv[1]
    if cmd == "sample" and len(argv) >= 5:
        old_pdf, new_pdf, out_dir = argv[2], argv[3], argv[4]
        print("[STATUS] see module docstring for which edition pairs are")
        print("confirmatory (the four minor pairs listed there) versus dev")
        print("data usable only for a mechanics check. This tool does not")
        print("know which one you passed - verify before reporting results.\n")
        r = stratified_sample(old_pdf, new_pdf)
        print(f"population by tier: {r['population_by_tier']}")
        print(f"drawn by tier:       {r['drawn_by_tier']}")
        print(f"sample weights:      {r['sample_weight_by_tier']}")
        print(f"total drawn:         {r['total_drawn']} (target {r['total_target']})")
        csv_path, ctx_path = write_annotation_packet(r, out_dir)
        print(f"\nwrote {csv_path}")
        print(f"wrote {ctx_path}")
        return 0
    if cmd == "kappa" and len(argv) >= 4:
        k = compute_kappa(argv[2], argv[3])
        print(json.dumps(k, indent=2))
        return 0
    print(__doc__)
    return 2


if __name__ == "__main__":
    sys.exit(main(sys.argv))
