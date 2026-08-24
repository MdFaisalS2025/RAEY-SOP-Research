"""
One-off driver: scores H3'a/H3'b/H3'c (PREREGISTRATION.md's H3' entries)
against the completed, blind H3' annotation.

AUDIT FIX (2026-08-23): previously pointed at E.xlsx/F.xlsx, the H3'
annotator pair later found to be the same duplication failure that hit
the main round's A/B pair (PREREGISTRATION.md's 2026-08-18 CRITICAL
CORRECTION) - their claimed "kappa=1.0000, 0/92 disagreements" was a
file agreeing with a copy of itself, not real ground truth. Now points
at Annotator_G_ANNOTATION.xlsx / Annotator_H_ANNOTATION.xlsx, a
genuinely independent pair (verified via run_h3prime_second_annotator.py:
different file hashes, different answers - Cohen's kappa 0.9414, 87/92
agree). The 5 genuine disagreements (all the same pattern: one annotator
said NONE, the other CANNOT_DETERMINE - a real interpretive question, not
noise) are excluded from ground truth pending real adjudication
(build_h3prime_adjudication.py generates the workbook) rather than
silently resolved - this script no longer raises on disagreement, it
scores the 87 resolved items and reports the 5 pending ones explicitly.

Per the user's explicit decision (PREREGISTRATION.md, the guideline-
boundary-bug entry): the 11 bullet-census items caught in the fresh
pair's "Delirium with HyperAgitation" boundary-bleed (FEASIBILITY.md
section 56) are excluded from scoring here, reported as a separate,
distinct finding rather than allowed to dilute the T2-fix result.

H3'a: v2 (fixed) vs v1 (original) accuracy, on the 21 clean bullet items.
H3'b: v2 (fixed) vs B2 accuracy, on the same 21 clean bullet items.
H3'c: v1(=v2, identical on ordinal items) vs B2 accuracy, on the 60
      ordinal items - an independent replication of the original H3's
      ordinal-item finding (FEASIBILITY.md section 54.1) on fresh data.

n=1 pair, so only item-level bootstrap is possible - stated as a scoping
fact in the pre-commitment entry, not a deviation.

Run:
    cd sop-guard/backend
    python -m app.research.cross_edition.annotation_packets.run_h3prime_test
"""
from __future__ import annotations

import csv
import json
import random
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[4]))  # backend/ on sys.path

from app.research.cross_edition.annotation import load_completed_xlsx, _norm_answer  # noqa: E402
from app.research.cross_edition.annotation_packets.run_full_comparison import (  # noqa: E402
    benjamini_hochberg,
)

BASE = Path(__file__).parent / "h3prime_tennessee_2022_2024"
ANNOTATOR_G = BASE / "Annotator_G_ANNOTATION.xlsx"
ANNOTATOR_H = BASE / "Annotator_H_ANNOTATION.xlsx"
ADJUDICATION_FILE = BASE / "H3prime_Adjudication_5_items_COMPLETED.xlsx"  # optional

CONTAMINATED_GUIDELINE = "Delirium with HyperAgitation"


def build_ground_truth() -> tuple[dict[str, str], list[str]]:
    """Returns (ground_truth, pending_sids). pending_sids lists items
    excluded from ground_truth because G/H disagreed and no completed
    adjudication file was found - reported explicitly, not silently
    dropped."""
    g = load_completed_xlsx(str(ANNOTATOR_G))
    h = load_completed_xlsx(str(ANNOTATOR_H))
    sheet_g = list(g.keys())[0]
    sheet_h = list(h.keys())[0]
    ga = {sid: _norm_answer(v["correspondence"]) for sid, v in g[sheet_g].items()}
    ha = {sid: _norm_answer(v["correspondence"]) for sid, v in h[sheet_h].items()}
    common = set(ga) & set(ha)

    gt = {}
    pending = []
    for sid in common:
        if ga[sid] == ha[sid]:
            gt[sid] = ga[sid]
        else:
            pending.append(sid)

    if pending and ADJUDICATION_FILE.exists():
        adj_answers = _load_h3prime_adjudication(ADJUDICATION_FILE)
        still_pending = []
        for sid in pending:
            if sid in adj_answers:
                gt[sid] = adj_answers[sid]
            else:
                still_pending.append(sid)
        pending = still_pending

    return gt, sorted(pending)


def _load_h3prime_adjudication(path: Path) -> dict[str, str]:
    """Reads build_h3prime_adjudication.py's own column layout directly
    (sample_id=col1, final_correspondence=col6) - NOT load_completed_xlsx,
    whose column layout (correspondence=col7) matches the original blind
    annotation packets, not this adjudication sheet's different layout.
    Matches run_h3_test.py's load_adjudicated(), which reads its own
    43-item adjudication sheet's actual columns the same way rather than
    reusing the blind-packet reader."""
    from openpyxl import load_workbook
    wb = load_workbook(path, data_only=True)
    ws = wb["Adjudicate these 5"]
    out = {}
    for r in range(2, ws.max_row + 1):
        sid = ws.cell(row=r, column=1).value
        final_corr = ws.cell(row=r, column=6).value
        if sid and final_corr:
            out[sid] = _norm_answer(final_corr)
    return out


def bootstrap_ci(pairs: list[tuple[int, int]], n_boot: int = 10000, seed: int = 20260818):
    """pairs: list of (a_correct, b_correct). Returns point estimate and
    95% CI for the paired difference (a - b), item-level resampling only
    (n=1 pair here, so no pair-level unit exists)."""
    n = len(pairs)
    point = sum(a - b for a, b in pairs) / n
    rng = random.Random(seed)
    diffs = []
    for _ in range(n_boot):
        sample = [pairs[rng.randrange(n)] for _ in range(n)]
        diffs.append(sum(a - b for a, b in sample) / n)
    diffs.sort()
    lo_i, hi_i = int(0.025 * n_boot), int(0.975 * n_boot) - 1
    # p-value for BH (2026-08-18 audit round 3, Phase 1: this was
    # pre-registered - PREREGISTRATION.md's H3' pre-commitment entry commits
    # "Benjamini-Hochberg applied across {H3'a, H3'b, H3'c} as its own
    # family" - but never applied. Null: diff<=0 (H3'a/b/c's shared
    # confirmation criterion requires diff>0). Percentile p, matching the
    # convention already used for H3/H4/H5 in run_full_comparison.py.
    p_value = sum(1 for d in diffs if d <= 0) / n_boot
    return {
        "n": n, "point_estimate": round(point, 4),
        "ci95_low": round(diffs[lo_i], 4), "ci95_high": round(diffs[hi_i], 4),
        "p_value": round(p_value, 4),
    }


def main():
    gt, pending = build_ground_truth()
    print(f"ground truth items: {len(gt)}")
    if pending:
        print(f"PENDING ADJUDICATION (excluded from ground truth, not scored): "
              f"{len(pending)} items - {pending}")
        print(f"  -> generate/fill build_h3prime_adjudication.py's workbook, "
              f"save as {ADJUDICATION_FILE.name}, and re-run to include them.")

    with open(BASE / "annotation_packet.csv", encoding="utf-8") as fh:
        packet = {r["sample_id"]: r for r in csv.DictReader(fh)}
    with open(BASE / "master_scoring.csv", encoding="utf-8") as fh:
        master = {r["sample_id"]: r for r in csv.DictReader(fh)}

    contaminated = {sid for sid, r in master.items()
                     if r["is_bullet_kind"] == "True"
                     and packet[sid]["old_guideline"] == CONTAMINATED_GUIDELINE}
    clean_bullet = [sid for sid, r in master.items()
                     if r["is_bullet_kind"] == "True" and sid not in contaminated]
    ordinal = [sid for sid, r in master.items() if r["is_bullet_kind"] != "True"]

    print(f"contaminated (excluded): {len(contaminated)}")
    print(f"clean bullet items: {len(clean_bullet)}")
    print(f"ordinal items: {len(ordinal)}")

    def scored(sids, pred_key_a, pred_key_b):
        pairs = []
        n_cannot_determine = 0
        for sid in sids:
            truth = gt.get(sid)
            if truth is None:
                continue
            if truth == "cannot_determine":
                n_cannot_determine += 1
                continue
            a_pred = _norm_answer(master[sid][pred_key_a])
            b_pred = _norm_answer(master[sid][pred_key_b])
            a_correct = 1 if a_pred == truth else 0
            b_correct = 1 if b_pred == truth else 0
            pairs.append((a_correct, b_correct))
        return pairs, n_cannot_determine

    h3a_pairs, h3a_cd = scored(clean_bullet, "v2_predicted_item_id", "v1_predicted_item_id")
    h3b_pairs, h3b_cd = scored(clean_bullet, "v2_predicted_item_id", "b2_predicted_item_id")
    h3c_pairs, h3c_cd = scored(ordinal, "v1_predicted_item_id", "b2_predicted_item_id")

    report = {}
    for name, pairs, cd, sids in [
        ("H3a_v2_vs_v1_bullet", h3a_pairs, h3a_cd, clean_bullet),
        ("H3b_v2_vs_B2_bullet", h3b_pairs, h3b_cd, clean_bullet),
        ("H3c_v1_vs_B2_ordinal", h3c_pairs, h3c_cd, ordinal),
    ]:
        boot = bootstrap_ci(pairs)
        a_acc = sum(a for a, _ in pairs) / len(pairs)
        b_acc = sum(b for _, b in pairs) / len(pairs)
        confirmed = boot["point_estimate"] > 0 and boot["ci95_low"] > 0
        report[name] = {
            "n_scored": len(pairs), "n_cannot_determine_excluded": cd,
            "n_total_population": len(sids),
            "a_accuracy": round(a_acc, 4), "b_accuracy": round(b_acc, 4),
            "bootstrap": boot, "confirmed": confirmed,
        }
        print(f"\n=== {name} ===")
        print(f"  n_scored={len(pairs)}  cannot_determine_excluded={cd}")
        print(f"  accuracy: a={a_acc:.4f}  b={b_acc:.4f}")
        print(f"  paired diff (a-b): {boot}")
        print(f"  CONFIRMED: {confirmed}")

    # BH across {H3'a, H3'b, H3'c} - pre-registered (PREREGISTRATION.md's
    # H3' pre-commitment entry) but never applied until this audit round.
    pvals = {name: report[name]["bootstrap"]["p_value"] for name in report}
    bh = benjamini_hochberg(pvals)
    report["benjamini_hochberg"] = bh
    report["_pending_adjudication"] = pending
    report["_ground_truth_source"] = (
        "Annotator_G_ANNOTATION.xlsx / Annotator_H_ANNOTATION.xlsx "
        "(genuinely independent, Cohen's kappa 0.9414 - see "
        "h3prime_second_annotator_report.json). Replaces the retracted "
        "E/F pair (2026-08-18 duplication finding)."
    )
    print(f"\nBH-adjusted p-values (H3'a/b/c): {bh}")
    print("Note: H3'a/H3'b's p=1.0 reflects the already-established UNTESTABLE "
          "finding (degenerate population, 20/21 items T1-tier where the fix "
          "is a no-op by construction) - not literal strong evidence against, "
          "consistent with FEASIBILITY.md's H3'a/H3'b write-up.")

    out_path = BASE / "h3prime_test_report.json"
    with open(out_path, "w", encoding="utf-8") as fh:
        json.dump(report, fh, indent=2)
    print(f"\nwrote {out_path}")


if __name__ == "__main__":
    main()
