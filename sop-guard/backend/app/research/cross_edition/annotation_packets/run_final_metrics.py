"""
One-off driver: assembles final, complete ground truth for all 240 sampled
items (197 from clear 4-rater majority + 43 from the completed adjudication
sheet) and computes PREREGISTRATION.md section 6's metrics against the
method's original predictions. Not part of the pipeline itself.
"""
import csv
import json
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[4]))  # backend/ on sys.path

from app.research.cross_edition.annotation import (  # noqa: E402
    load_completed_xlsx, _norm_answer, majority_vote, compute_section6_metrics,
)

BASE = Path(__file__).parent

ANNOTATOR_FILES = {
    "A": r"C:\Users\Faisal\Desktop\Annotator_A.xlsx",
    "B": r"C:\Users\Faisal\Desktop\Annotator_B.xlsx",
    "C": r"C:\Users\Faisal\Desktop\Annotator_C.xlsx",
    "D": r"C:\Users\Faisal\Desktop\Annotator_D.xlsx",
}
ADJUDICATION_FILE = r"C:\Users\Faisal\Downloads\Adjudication_43_items_completed.xlsx"

PAIR_SLUGS = {
    "Tennessee 2017→2018": "tennessee_2017_2018",
    "Pennsylvania 2021→2023": "pennsylvania_2021_2023",
    "Connecticut 2022.1→2023.1": "connecticut_v20221_v20231",
    "Connecticut 2023.1→2024.1": "connecticut_v20231_v20241",
}


def load_adjudicated() -> dict[str, dict[str, str]]:
    from openpyxl import load_workbook
    wb = load_workbook(ADJUDICATION_FILE, data_only=True)
    ws = wb["Adjudicate these 43"]
    out: dict[str, dict[str, str]] = {p: {} for p in PAIR_SLUGS}
    for r in range(2, ws.max_row + 1):
        pair = ws.cell(row=r, column=1).value
        sid = ws.cell(row=r, column=2).value
        final_corr = ws.cell(row=r, column=9).value
        if pair and sid:
            out[pair][sid] = final_corr
    return out


def build_ground_truth() -> dict[str, dict[str, str]]:
    raw = {label: load_completed_xlsx(path) for label, path in ANNOTATOR_FILES.items()}
    adjudicated = load_adjudicated()

    ground_truth: dict[str, dict[str, str]] = {}
    for pair in PAIR_SLUGS:
        norm = {
            label: {sid: _norm_answer(v["correspondence"])
                    for sid, v in raw[label][pair].items()}
            for label in ANNOTATOR_FILES
        }
        mv = majority_vote([norm["A"], norm["B"], norm["C"], norm["D"]])
        gt = {}
        for sid, result in mv.items():
            if result["answer"] is not None:
                gt[sid] = result["answer"]
            else:
                adj = adjudicated[pair].get(sid)
                if adj is None:
                    raise ValueError(f"missing adjudication for {pair} / {sid}")
                gt[sid] = _norm_answer(adj)
        ground_truth[pair] = gt
    return ground_truth


def load_method_rows(slug: str) -> list[dict]:
    with open(BASE / slug / "annotation_packet.csv", encoding="utf-8") as f:
        return list(csv.DictReader(f))


def main():
    ground_truth = build_ground_truth()

    report = {"pairs": {}, "pooled": {}}
    pooled_gt: dict[str, str] = {}
    pooled_rows: list[dict] = []

    for pair, slug in PAIR_SLUGS.items():
        rows = load_method_rows(slug)
        metrics = compute_section6_metrics(ground_truth[pair], rows)
        report["pairs"][pair] = metrics

        for sid, ans in ground_truth[pair].items():
            key = f"{pair}::{sid}"
            pooled_gt[key] = ans
        for r in rows:
            r2 = dict(r)
            r2["sample_id"] = f"{pair}::{r['sample_id']}"
            pooled_rows.append(r2)

        print(f"=== {pair} ===")
        print(json.dumps(metrics, indent=2))
        print()

    pooled_metrics = compute_section6_metrics(pooled_gt, pooled_rows)
    report["pooled"] = pooled_metrics
    print("=== POOLED (all 4 pairs, 240 items) ===")
    print(json.dumps(pooled_metrics, indent=2))

    out_path = BASE / "section6_final_metrics.json"
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(report, f, indent=2, ensure_ascii=False)
    print(f"\nwrote {out_path}")


if __name__ == "__main__":
    main()
