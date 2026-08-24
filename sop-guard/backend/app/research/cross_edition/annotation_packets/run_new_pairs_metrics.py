"""
One-off driver: assembles ground truth for the two new confirmatory
pairs (PREREGISTRATION.md's 2026-08-23 "Fifth and sixth confirmatory
pairs" entry - Tennessee Sept2024->09.11.2025, Connecticut
v2024.1->v2025.1) from the completed Annotator I / Annotator J
workbooks, and computes section 6's metrics against the method's
predictions, mirroring run_final_metrics.py's structure for the
original four pairs.

Two raters only (I, J), not four - majority_vote's >n/2 rule reduces
to "both agree" for n=2, so any I/J disagreement is correctly flagged
needs_adjudication rather than silently resolved, matching the
2-rater treatment already established for H3' (run_h3prime_test.py).

Independence verified first (byte-hash + cell-level answer comparison,
run_boundary_scoring.py's established pattern) before any metric is
trusted, per section 10's standing rule and the main round's A/B
duplication precedent.

Not part of the pipeline itself - a report generator over already-
collected annotation data. Does not modify item_align.py,
item_parser.py, corpus_probe.py, or edition_align.py.

Run:
    cd sop-guard/backend
    python -m app.research.cross_edition.annotation_packets.run_new_pairs_metrics
"""
from __future__ import annotations

import csv
import hashlib
import json
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[4]))  # backend/ on sys.path

from app.research.cross_edition.annotation import (  # noqa: E402
    load_completed_xlsx, _norm_answer, majority_vote, compute_section6_metrics,
)
from app.research.cross_edition.annotation_packets.run_4rater_analysis import cohens_kappa  # noqa: E402

BASE = Path(__file__).parent

ANNOTATOR_I = BASE / "Annotator_I_ANNOTATION.xlsx"
ANNOTATOR_J = BASE / "Annotator_J_ANNOTATION.xlsx"
ADJUDICATION_FILE = BASE / "NewPairs_Adjudication_COMPLETED.xlsx"  # optional, generated on demand

PAIR_SLUGS = {
    "Tennessee Sept2024→09.11.2025 (": "tennessee_sept2024_20250911",
    "Connecticut v2024.1→v2025.1 (ne": "connecticut_v20241_v20251",
}


def verify_independence(path_i: Path, path_j: Path) -> dict:
    """Mirrors run_boundary_scoring.py's / run_h3prime_second_annotator.py's
    verify_independence exactly. Raises rather than silently proceeding on
    a suspicious match."""
    hash_i = hashlib.md5(path_i.read_bytes()).hexdigest()
    hash_j = hashlib.md5(path_j.read_bytes()).hexdigest()
    identical_bytes = hash_i == hash_j

    raw_i = load_completed_xlsx(str(path_i))
    raw_j = load_completed_xlsx(str(path_j))
    per_pair = {}
    all_identical = True
    for sheet in PAIR_SLUGS:
        ans_i = {sid: _norm_answer(v["correspondence"]) for sid, v in raw_i.get(sheet, {}).items()}
        ans_j = {sid: _norm_answer(v["correspondence"]) for sid, v in raw_j.get(sheet, {}).items()}
        identical = ans_i == ans_j and len(ans_i) > 0
        per_pair[sheet] = {"n_i": len(ans_i), "n_j": len(ans_j), "identical_answers": identical}
        all_identical = all_identical and identical

    report = {
        "hash_i": hash_i, "hash_j": hash_j, "identical_bytes": identical_bytes,
        "per_pair": per_pair, "identical_answers_all_pairs": all_identical,
    }
    print(f"Independence check: {json.dumps(report, indent=2)}")
    if identical_bytes:
        raise RuntimeError("Annotator I/J files are byte-identical - collection "
                            "failure, not agreement. Do not proceed.")
    if all_identical:
        raise RuntimeError("Annotator I/J answers are identical on every pair despite "
                            "different file bytes - the same pattern that caught the "
                            "original E/F and A/B duplications. Treat as unverified "
                            "until re-collected.")
    return report


def _load_new_pairs_adjudication(path: Path) -> dict[str, dict[str, str]]:
    """Reads a from-scratch adjudication workbook's own column layout
    directly via openpyxl - NOT load_completed_xlsx, whose layout matches
    the blind annotation packets, not an adjudication sheet. Expected
    columns: A=pair_sheet_title, B=sample_id, ... final_correspondence in
    the last column (mirrors run_final_metrics.py's load_adjudicated())."""
    from openpyxl import load_workbook
    wb = load_workbook(path, data_only=True)
    ws = wb.worksheets[-1]  # data sheet, not "READ ME FIRST"
    out: dict[str, dict[str, str]] = {p: {} for p in PAIR_SLUGS}
    header = [c.value for c in ws[1]]
    final_col = None
    for idx, h in enumerate(header, start=1):
        if h and "final" in str(h).lower() and "correspond" in str(h).lower():
            final_col = idx
    if final_col is None:
        final_col = ws.max_column
    for r in range(2, ws.max_row + 1):
        pair = ws.cell(row=r, column=1).value
        sid = ws.cell(row=r, column=2).value
        final_corr = ws.cell(row=r, column=final_col).value
        if pair in out and sid and final_corr:
            out[pair][sid] = _norm_answer(final_corr)
    return out


def build_ground_truth() -> tuple[dict[str, dict[str, str]], dict[str, list[str]]]:
    """Returns ({sheet: {sample_id: answer}}, {sheet: [pending sample_ids]}) -
    pending items are excluded from ground truth, not silently resolved,
    matching run_h3prime_test.py's established pattern for 2-rater data."""
    raw_i = load_completed_xlsx(str(ANNOTATOR_I))
    raw_j = load_completed_xlsx(str(ANNOTATOR_J))

    adjudicated = {}
    if ADJUDICATION_FILE.exists():
        adjudicated = _load_new_pairs_adjudication(ADJUDICATION_FILE)

    ground_truth: dict[str, dict[str, str]] = {}
    pending: dict[str, list[str]] = {}
    for sheet in PAIR_SLUGS:
        norm_i = {sid: _norm_answer(v["correspondence"]) for sid, v in raw_i[sheet].items()}
        norm_j = {sid: _norm_answer(v["correspondence"]) for sid, v in raw_j[sheet].items()}
        mv = majority_vote([norm_i, norm_j])
        gt, still_pending = {}, []
        for sid, result in mv.items():
            if result["answer"] is not None:
                gt[sid] = result["answer"]
            else:
                adj = adjudicated.get(sheet, {}).get(sid)
                if adj is not None:
                    gt[sid] = adj
                else:
                    still_pending.append(sid)
        ground_truth[sheet] = gt
        pending[sheet] = sorted(still_pending)
    return ground_truth, pending


def load_method_rows(slug: str) -> list[dict]:
    with open(BASE / slug / "annotation_packet.csv", encoding="utf-8") as f:
        return list(csv.DictReader(f))


def main():
    if not ANNOTATOR_I.exists() or not ANNOTATOR_J.exists():
        print(f"Waiting for both files to exist at:\n  {ANNOTATOR_I}\n  {ANNOTATOR_J}")
        return

    verify_independence(ANNOTATOR_I, ANNOTATOR_J)

    raw_i = load_completed_xlsx(str(ANNOTATOR_I))
    raw_j = load_completed_xlsx(str(ANNOTATOR_J))

    ground_truth, pending = build_ground_truth()

    report: dict = {"pairs": {}, "pooled": {}, "kappa": {}, "pending_adjudication": pending}
    pooled_gt: dict[str, str] = {}
    pooled_rows: list[dict] = []
    pooled_i: dict[str, str] = {}
    pooled_j: dict[str, str] = {}

    for sheet, slug in PAIR_SLUGS.items():
        norm_i = {sid: _norm_answer(v["correspondence"]) for sid, v in raw_i[sheet].items()}
        norm_j = {sid: _norm_answer(v["correspondence"]) for sid, v in raw_j[sheet].items()}
        ck = cohens_kappa(norm_i, norm_j)
        report["kappa"][sheet] = ck
        print(f"=== {sheet} ===")
        print(f"  Cohen's kappa (I/J): {ck['cohens_kappa']}  "
              f"(observed agreement {ck['observed_agreement']}, "
              f"{len(ck['disagreements'])}/{ck['n']} disagreements)")
        if pending[sheet]:
            print(f"  PENDING ADJUDICATION (excluded from ground truth): "
                  f"{len(pending[sheet])} items - {pending[sheet]}")

        rows = load_method_rows(slug)
        metrics = compute_section6_metrics(ground_truth[sheet], rows)
        report["pairs"][sheet] = metrics
        print(json.dumps(metrics, indent=2))
        print()

        for sid, ans in ground_truth[sheet].items():
            key = f"{sheet}::{sid}"
            pooled_gt[key] = ans
        for r in rows:
            r2 = dict(r)
            r2["sample_id"] = f"{sheet}::{r['sample_id']}"
            pooled_rows.append(r2)
        for sid, v in norm_i.items():
            pooled_i[f"{sheet}::{sid}"] = v
        for sid, v in norm_j.items():
            pooled_j[f"{sheet}::{sid}"] = v

    pooled_ck = cohens_kappa(pooled_i, pooled_j)
    report["kappa"]["pooled"] = pooled_ck
    pooled_metrics = compute_section6_metrics(pooled_gt, pooled_rows)
    report["pooled"] = pooled_metrics

    print("=== POOLED (both new pairs, 120 items) ===")
    print(f"  Cohen's kappa (I/J, pooled): {pooled_ck['cohens_kappa']}  "
          f"(observed agreement {pooled_ck['observed_agreement']}, "
          f"{len(pooled_ck['disagreements'])}/{pooled_ck['n']} disagreements)")
    print(json.dumps(pooled_metrics, indent=2))

    out_path = BASE / "new_pairs_final_metrics.json"
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(report, f, indent=2, ensure_ascii=False)
    print(f"\nwrote {out_path}")


if __name__ == "__main__":
    main()
