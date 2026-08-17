"""
One-off script: builds a single, self-contained Excel workbook per annotator,
combining the blind CSV + context JSON for all 4 confirmatory pairs into one
file each, with a plain-language instructions tab, the new-edition guideline
text embedded directly (no separate JSON lookup needed), and the columns the
annotator must fill in clearly highlighted.

Not part of the research pipeline itself - a one-time deliverable-formatting
step, run once, output committed alongside the source CSV/JSON it was built
from.
"""
import csv
import json
import re
from pathlib import Path

# XML 1.0 only allows #x9 | #xA | #xD | [#x20-#xD7FF] | [#xE000-#xFFFD] |
# [#x10000-#x10FFFF]. PDF text extraction sometimes leaves stray control
# characters (from bullet-glyph private-use-area remnants or extraction
# noise) that openpyxl's writer rejects outright. Strip anything outside
# the legal range rather than let the whole workbook build fail on one row.
_ILLEGAL_XML_CHARS_RE = re.compile(
    "[\x00-\x08\x0b\x0c\x0e-\x1f\x7f-\x84\x86-\x9f\ud800-\udfff﷐-﷯￾￿]"
)


def sanitize(text: str) -> str:
    if not isinstance(text, str):
        return text
    return _ILLEGAL_XML_CHARS_RE.sub("", text)

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

BASE = Path(__file__).parent

PAIRS = [
    ("tennessee_2017_2018", "Tennessee 2017→2018"),
    ("pennsylvania_2021_2023", "Pennsylvania 2021→2023"),
    ("connecticut_v20221_v20231", "Connecticut 2022.1→2023.1"),
    ("connecticut_v20231_v20241", "Connecticut 2023.1→2024.1"),
]

FONT_NAME = "Calibri"
HEADER_FILL = PatternFill("solid", fgColor="4472C4")
HEADER_FONT = Font(name=FONT_NAME, bold=True, color="FFFFFF", size=11)
FILLIN_FILL = PatternFill("solid", fgColor="FFF2CC")  # soft yellow
FILLIN_FONT = Font(name=FONT_NAME, size=11)
NORMAL_FONT = Font(name=FONT_NAME, size=11)
WRAP_TOP = Alignment(wrap_text=True, vertical="top")
THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

RELATION_OPTIONS = "unchanged,reworded,substantive,merged,split,moved"

COLUMNS = [
    # (header, width, key)
    ("sample_id", 9, "sample_id"),
    ("old_item_id", 26, "old_item_id"),
    ("old_guideline", 22, "old_guideline"),
    ("old_section", 16, "old_section"),
    ("OLD RECOMMENDATION (read this)", 46, "old_text"),
    ("NEW EDITION - everything in the matching guideline (search this for a match)", 60, "new_guideline_text"),
    ("YOUR ANSWER: matching new item ID, or NONE, or CANNOT_DETERMINE", 30, "annotator_correspondence"),
    ("YOUR ANSWER: relation", 16, "annotator_relation"),
    ("YOUR NOTES (optional)", 30, "annotator_notes"),
]
FILLIN_KEYS = {"annotator_correspondence", "annotator_relation", "annotator_notes"}

INSTRUCTIONS = [
    ("What you're doing", None),
    (None, "You're checking whether specific safety recommendations from an older "
           "edition of an EMS (emergency medical) protocol document still exist in "
           "a newer edition - and if so, whether they changed."),
    (None, "This is NOT a medical judgement. It's a matching task: 'is this the same "
           "instruction, somewhere in the new document - reworded or not - or was it "
           "removed?'"),
    ("Where to work", None),
    (None, "This workbook has one tab per document pair (see tabs at the bottom). "
           "Each tab has 60 rows to judge. Work through all rows on all tabs."),
    ("What to read, per row", None),
    (None, "Column E (yellow header) - the OLD recommendation."),
    (None, "Column F - the FULL new edition guideline it might now belong to. Read "
           "through it and look for the same instruction."),
    ("What to fill in (yellow cells, columns G-I)", None),
    (None, "Column G - type the matching item's ID exactly as it appears at the start "
           "of that item in column F, OR type NONE if you're confident it was removed, "
           "OR type CANNOT_DETERMINE if you truly can't tell."),
    (None, "Column H - pick one from the dropdown: unchanged / reworded / substantive "
           "/ merged / split / moved. (Leave blank if you answered NONE.)"),
    (None, "Column I - optional. Explain anything unsure, especially CANNOT_DETERMINE."),
    ("Rules", None),
    (None, "1. Work alone. Don't discuss rows with the other annotator, or compare "
           "answers, until you have BOTH finished every tab. This is the whole point - "
           "we're measuring how often two people agree without coordinating."),
    (None, "2. Don't guess if you're not sure - CANNOT_DETERMINE is a normal, expected "
           "answer sometimes, not a failure."),
    (None, "3. You do not need any other file. Everything you need is in column F."),
    ("When you're done", None),
    (None, "Save this file and send it back exactly as-is (don't rename tabs or "
           "columns). That's it."),
]


def load_pair(slug: str):
    csv_path = BASE / slug / "annotation_packet.csv"
    ctx_path = BASE / slug / "annotation_context.json"
    with open(csv_path, encoding="utf-8") as f:
        rows = list(csv.DictReader(f))
    with open(ctx_path, encoding="utf-8") as f:
        ctx = json.load(f)
    return rows, ctx


def render_new_guideline(ctx_entry: dict) -> str:
    items = ctx_entry.get("new_guideline_full_text", [])
    if not items:
        return "(no items found in this guideline in the new edition)"
    blocks = []
    for it in items:
        blocks.append(f"[{it['item_id']}]\n{it['text']}")
    return sanitize("\n\n".join(blocks))


def build_instructions_sheet(wb: Workbook, annotator_label: str):
    ws = wb.active
    ws.title = "READ ME FIRST"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 100

    r = 1
    title_cell = ws.cell(row=r, column=1, value=f"Annotation task - {annotator_label}")
    title_cell.font = Font(name=FONT_NAME, bold=True, size=16, color="1F4E78")
    r += 2

    for heading, body in INSTRUCTIONS:
        if heading is not None:
            c = ws.cell(row=r, column=1, value=heading)
            c.font = Font(name=FONT_NAME, bold=True, size=13, color="1F4E78")
            r += 1
        if body is not None:
            c = ws.cell(row=r, column=1, value=body)
            c.font = Font(name=FONT_NAME, size=11.5)
            c.alignment = Alignment(wrap_text=True, vertical="top")
            ws.row_dimensions[r].height = 32
            r += 1
    r += 1
    note = ws.cell(
        row=r, column=1,
        value="Tabs (bottom of screen): each is one document pair, 60 rows each.",
    )
    note.font = Font(name=FONT_NAME, italic=True, size=11, color="595959")


def build_pair_sheet(wb: Workbook, sheet_title: str, rows: list[dict], ctx: dict):
    ws = wb.create_sheet(title=sheet_title[:31])
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"

    for col_idx, (header, width, _key) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = WRAP_TOP
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[1].height = 30

    dv = DataValidation(
        type="list", formula1=f'"{RELATION_OPTIONS}"', allow_blank=True,
        showDropDown=False,
    )
    dv.error = "Pick one from the list."
    dv.errorTitle = "Invalid entry"
    ws.add_data_validation(dv)

    for i, row in enumerate(rows):
        r = i + 2
        sid = row["sample_id"]
        ctx_entry = ctx.get(sid, {})
        values = {
            "sample_id": sanitize(sid),
            "old_item_id": sanitize(row["old_item_id"]),
            "old_guideline": sanitize(row["old_guideline"]),
            "old_section": sanitize(row["old_section"]),
            "old_text": sanitize(row["old_text"]),
            "new_guideline_text": render_new_guideline(ctx_entry),
            "annotator_correspondence": "",
            "annotator_relation": "",
            "annotator_notes": "",
        }
        for col_idx, (_header, _width, key) in enumerate(COLUMNS, start=1):
            cell = ws.cell(row=r, column=col_idx, value=values[key])
            cell.border = BORDER
            cell.alignment = WRAP_TOP
            if key in FILLIN_KEYS:
                cell.fill = FILLIN_FILL
                cell.font = FILLIN_FONT
            else:
                cell.font = NORMAL_FONT
        ws.row_dimensions[r].height = 60
        dv.add(ws.cell(row=r, column=8))  # column H = annotator_relation

    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{len(rows) + 1}"


def build_workbook(annotator_key: str, annotator_label: str, out_path: Path):
    wb = Workbook()
    build_instructions_sheet(wb, annotator_label)
    for slug, title in PAIRS:
        rows, ctx = load_pair(slug)
        build_pair_sheet(wb, title, rows, ctx)
    wb.save(out_path)
    print(f"wrote {out_path}")


if __name__ == "__main__":
    build_workbook(
        "A", "Annotator A",
        BASE / "Annotator_A" / "Annotator_A_ANNOTATION.xlsx",
    )
    build_workbook(
        "B", "Annotator B",
        BASE / "Annotator_B" / "Annotator_B_ANNOTATION.xlsx",
    )
