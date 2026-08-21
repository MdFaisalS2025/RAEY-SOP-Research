"""
Scores the boundary-annotation workbooks (Workstream A, novelty-audit
plan) once returned: measures item_parser.py's real guideline-boundary
detection precision/recall/F1 against hand-annotated ground truth, for
each of Tennessee's four editions.

MANDATORY FIRST STEP (per the plan, after the main round's annotator-
duplication failure): verify the two returned files are NOT identical -
by file hash AND by cell-level list comparison - before computing
anything else. A duplicated pair must be treated as a collection
failure, not agreement, exactly as corrected in PREREGISTRATION.md's
2026-08-18 CRITICAL CORRECTION entry.

Matching method: reuses item_align.match_guidelines (title token-overlap,
floor 0.5, greedy) COMPLETELY UNCHANGED to align the annotator's ordered
title list against item_parser.parse(pdf).guidelines. This is not a new
matching algorithm invented for this check - it is the same function
already used and validated throughout this study for cross-edition
guideline correspondence, applied here to align two title lists instead
of two editions' guidelines. Its greedy one-to-one consumption is exactly
the right behaviour for this measurement: if N real protocols were all
merged into one oversized parser guideline (section 56's mechanism), only
one annotator title can match that one parser guideline; the other N-1
correctly register as recall failures (missed boundaries), which is
precisely what should happen.

  Recall    = matched annotator titles / total annotator titles
              (fraction of REAL boundaries the parser detected as their
              own guideline)
  Precision = matched parser guidelines / total parser guidelines
              (fraction of DETECTED guidelines that correspond to a real,
              annotator-confirmed boundary)

Does not modify item_align.py or item_parser.py - imports match_guidelines
and parse() unchanged.

Run:
    cd sop-guard/backend
    python -m app.research.cross_edition.annotation_packets.run_boundary_scoring
"""
from __future__ import annotations

import hashlib
import json
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[4]))  # backend/ on sys.path

from openpyxl import load_workbook  # noqa: E402
from app.research.cross_edition.item_parser import parse  # noqa: E402
from app.research.cross_edition.item_align import match_guidelines  # noqa: E402

SP = r"C:\Users\Faisal\AppData\Local\Temp\claude\C--Users-Faisal-Desktop-research-paper\1642f160-3dba-4100-baa8-850fde74b388\scratchpad\protocols"
EDITIONS = [
    ("tn_2017", "Tennessee 2017 (Rev 11.7.2017)", f"{SP}\\tn_2017.pdf"),
    ("tn_2018", "Tennessee 2018 (Rev 7.7.18)", f"{SP}\\tn_2018.pdf"),
    ("tn_2022_23", "Tennessee 2022-23", f"{SP}\\tn_2022_23.pdf"),
    ("tn_sept2024", "Tennessee Sept2024 (2024-2025)", f"{SP}\\tn_sept2024.pdf"),
]

BASE = Path(__file__).parent / "boundary_annotation"
ANNOTATOR_1 = BASE / "Boundary_Annotator_1_COMPLETED.xlsx"
ANNOTATOR_2 = BASE / "Boundary_Annotator_2_COMPLETED.xlsx"


def load_titles(path: Path, sheet_title: str) -> list[str]:
    wb = load_workbook(path, data_only=True)
    ws = wb[sheet_title[:31]]
    titles = []
    for r in range(2, ws.max_row + 1):
        v = ws.cell(row=r, column=2).value  # column B = title
        if v and str(v).strip():
            titles.append(str(v).strip())
    return titles


def verify_independence(path_a: Path, path_b: Path) -> dict:
    """Mandatory pre-check. Returns a report; raises if the files are
    suspiciously identical, mirroring the discipline that should have
    caught the main round's annotator duplication the first time."""
    hash_a = hashlib.md5(path_a.read_bytes()).hexdigest()
    hash_b = hashlib.md5(path_b.read_bytes()).hexdigest()
    identical_bytes = hash_a == hash_b

    identical_lists = 0
    total_editions = 0
    for slug, title, _pdf in EDITIONS:
        ta = load_titles(path_a, title)
        tb = load_titles(path_b, title)
        total_editions += 1
        if ta == tb:
            identical_lists += 1

    report = {
        "hash_a": hash_a, "hash_b": hash_b, "identical_bytes": identical_bytes,
        "editions_with_identical_title_lists": identical_lists,
        "total_editions": total_editions,
    }
    print(f"Independence check: {report}")
    if identical_bytes:
        raise RuntimeError("Annotator files are byte-identical - collection failure, not agreement.")
    if identical_lists == total_editions:
        raise RuntimeError(
            "Every edition's title list is identical between annotators despite "
            "different file bytes - the same pattern already caught once this "
            "study (main round A/B, H3' E/F). Treat as unverified until re-collected."
        )
    return report


def score_edition(annotator_titles: list[str], pdf_path: str) -> dict:
    parser_titles = parse(pdf_path).guidelines
    # match_guidelines expects (old_titles, new_titles) - here "old" is the
    # human ground truth and "new" is the parser's output, an arbitrary but
    # consistent choice; the function is symmetric in structure.
    mapping = match_guidelines(annotator_titles, parser_titles)
    matched_annotator = len(mapping)
    matched_parser_titles = set(mapping.values())

    recall = matched_annotator / max(1, len(annotator_titles))
    precision = len(matched_parser_titles) / max(1, len(parser_titles))
    f1 = 2 * precision * recall / max(1e-9, precision + recall)

    missed = [t for t in annotator_titles if t not in mapping]
    spurious = [t for t in parser_titles if t not in matched_parser_titles]

    return {
        "n_annotator_titles": len(annotator_titles),
        "n_parser_guidelines": len(parser_titles),
        "n_matched": matched_annotator,
        "recall": round(recall, 4), "precision": round(precision, 4),
        "f1": round(f1, 4),
        "missed_boundaries": missed[:20],
        "spurious_guidelines": spurious[:20],
    }


def main():
    if not ANNOTATOR_1.exists() or not ANNOTATOR_2.exists():
        print(f"Waiting for completed files at:\n  {ANNOTATOR_1}\n  {ANNOTATOR_2}")
        return

    verify_independence(ANNOTATOR_1, ANNOTATOR_2)

    report: dict = {"per_edition": {}, "per_annotator": {"1": {}, "2": {}}}
    for annotator_key, path in [("1", ANNOTATOR_1), ("2", ANNOTATOR_2)]:
        for slug, title, pdf in EDITIONS:
            titles = load_titles(path, title)
            result = score_edition(titles, pdf)
            report["per_annotator"][annotator_key][slug] = result
            print(f"annotator {annotator_key} / {slug}: "
                  f"P={result['precision']:.4f} R={result['recall']:.4f} F1={result['f1']:.4f}")

    # Pooled F1 per edition, averaging the two annotators (a simple,
    # transparent combination rule fixed here rather than picking whichever
    # annotator's number looks better after seeing them).
    for slug, title, pdf in EDITIONS:
        f1s = [report["per_annotator"][k][slug]["f1"] for k in ("1", "2")]
        report["per_edition"][slug] = {"mean_f1": round(sum(f1s) / len(f1s), 4)}

    out = BASE / "boundary_scoring_report.json"
    with open(out, "w", encoding="utf-8") as f:
        json.dump(report, f, indent=2, ensure_ascii=False)
    print(f"\nwrote {out}")


if __name__ == "__main__":
    main()
