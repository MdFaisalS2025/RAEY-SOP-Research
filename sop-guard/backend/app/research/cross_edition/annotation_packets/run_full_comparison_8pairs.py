"""
One-off driver: formal re-run of H3/H4/H5 against the expanded 8-pair,
480-item confirmatory dataset - the follow-up flagged as outstanding when
the seventh/eighth (Massachusetts) pairs were committed
(PREREGISTRATION.md's 2026-08-24 "Seventh and eighth confirmatory pairs"
entry): "H3/H4/H5 and every other already-reported figure stand exactly
as reported on the 6-pair, 360-item dataset until a formal re-run against
the expanded 8-pair, 480-item dataset is separately committed and
performed."

Extends run_full_comparison_6pairs.py's exact method to the two
Massachusetts pairs, whose items require `item_parser_ma.py` (new, not
frozen) instead of the frozen item_parser.parse. Because the baseline
modules (baseline_b1_b3_b4.py, baseline_b2.py) each bind their own
module-level `parse` name at import time (`from item_parser import
parse`), monkeypatching item_parser.parse alone would NOT reach them -
each module's own `parse` attribute must be patched directly, the same
principle vlm_rescue_attempt.py already established for item_align.py.

PRELIMINARY: 2 of 120 Massachusetts items are still pending human
adjudication (MAPairs_Adjudication_2_items.xlsx, generated, awaiting
completion) - excluded from ground truth here, not guessed at, matching
every prior pending-adjudication treatment in this study.

Reports the original 6-pair pooled result (recomputed here, unchanged in
method), the Massachusetts-only 2-pair breakout, and the full 8-pair
pooled result.

Does not modify item_align.py, item_parser.py, corpus_probe.py, or
edition_align.py.

Run:
    cd sop-guard/backend
    python -m app.research.cross_edition.annotation_packets.run_full_comparison_8pairs
"""
from __future__ import annotations

import json
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[4]))  # backend/ on sys.path

from app.research.cross_edition.annotation import (  # noqa: E402
    load_completed_xlsx, _norm_answer, majority_vote,
)
from app.research.cross_edition import item_align, item_parser  # noqa: E402
from app.research.cross_edition import baseline_b1_b3_b4, baseline_b2  # noqa: E402
from app.research.cross_edition.item_parser_ma import parse as parse_ma  # noqa: E402
from app.research.cross_edition.annotation_packets.run_h3_test import (  # noqa: E402
    PAIRS as ORIGINAL_PAIRS, build_ground_truth as build_original_ground_truth,
)
from app.research.cross_edition.annotation_packets.run_full_comparison_6pairs import (  # noqa: E402
    NEW_PAIRS, build_new_pairs_ground_truth, build_records_for, run_hypothesis_tests,
)

BASE = Path(__file__).parent
SP = r"C:\Users\Faisal\Desktop\Hospital SOP's Research\corpus\protocols"

MA_PAIRS = {
    "Massachusetts v2025.1→v2026.1 (": (
        "massachusetts_v20251_v20261", f"{SP}\\ma_2025_fresh.pdf", f"{SP}\\ma_v20261.pdf"),
    "Massachusetts v2026.1→v2026.2 (": (
        "massachusetts_v20261_v20262", f"{SP}\\ma_v20261.pdf", f"{SP}\\ma_v20262.pdf"),
}

ANNOTATOR_K = BASE / "Annotator_K_ANNOTATION.xlsx"
ANNOTATOR_L = BASE / "Annotator_L_ANNOTATION.xlsx"
ADJUDICATION_FILE = BASE / "MAPairs_Adjudication_2_items_COMPLETED.xlsx"


def _load_ma_adjudication() -> dict[str, dict[str, str]]:
    out: dict[str, dict[str, str]] = {p: {} for p in MA_PAIRS}
    if not ADJUDICATION_FILE.exists():
        return out
    from openpyxl import load_workbook
    wb = load_workbook(ADJUDICATION_FILE, data_only=True)
    ws = wb.worksheets[-1]
    for r in range(2, ws.max_row + 1):
        pair = ws.cell(row=r, column=1).value
        sid = ws.cell(row=r, column=2).value
        final_corr = ws.cell(row=r, column=7).value
        if pair in out and sid and final_corr:
            out[pair][sid] = _norm_answer(final_corr)
    return out


def build_ma_ground_truth() -> tuple[dict[str, dict[str, str]], dict[str, list[str]]]:
    raw_k = load_completed_xlsx(str(ANNOTATOR_K))
    raw_l = load_completed_xlsx(str(ANNOTATOR_L))
    adjudicated = _load_ma_adjudication()

    ground_truth: dict[str, dict[str, str]] = {}
    pending: dict[str, list[str]] = {}
    for pair in MA_PAIRS:
        norm_k = {sid: _norm_answer(v["correspondence"]) for sid, v in raw_k[pair].items()}
        norm_l = {sid: _norm_answer(v["correspondence"]) for sid, v in raw_l[pair].items()}
        mv = majority_vote([norm_k, norm_l])
        gt, still_pending = {}, []
        for sid, result in mv.items():
            if result["answer"] is not None:
                gt[sid] = result["answer"]
            elif sid in adjudicated[pair]:
                gt[sid] = adjudicated[pair][sid]
            else:
                still_pending.append(sid)
        ground_truth[pair] = gt
        pending[pair] = sorted(still_pending)
    return ground_truth, pending


def build_ma_records(ground_truth: dict, start_idx: int) -> list[dict]:
    """Massachusetts needs item_parser_ma.parse, not the frozen parser -
    monkeypatch every module that bound its own `parse` name at import
    time (item_align, baseline_b1_b3_b4, baseline_b2), restore all three
    immediately after, regardless of outcome."""
    orig = {
        "item_align": item_align.parse,
        "item_parser": item_parser.parse,
        "b134": baseline_b1_b3_b4.parse,
        "b2": baseline_b2.parse,
    }
    item_align.parse = parse_ma
    item_parser.parse = parse_ma
    baseline_b1_b3_b4.parse = parse_ma
    baseline_b2.parse = parse_ma
    try:
        return build_records_for(MA_PAIRS, ground_truth, start_idx=start_idx)
    finally:
        item_align.parse = orig["item_align"]
        item_parser.parse = orig["item_parser"]
        baseline_b1_b3_b4.parse = orig["b134"]
        baseline_b2.parse = orig["b2"]


def main():
    original_gt = build_original_ground_truth()
    new_gt, new_pending = build_new_pairs_ground_truth()
    ma_gt, ma_pending = build_ma_ground_truth()

    total_pending = sum(len(v) for v in ma_pending.values())
    print(f"MA items pending adjudication (excluded below): {total_pending} - {ma_pending}")

    original_records = build_records_for(ORIGINAL_PAIRS, original_gt, start_idx=0)
    new_records = build_records_for(NEW_PAIRS, new_gt, start_idx=len(ORIGINAL_PAIRS))
    ma_records = build_ma_records(ma_gt, start_idx=len(ORIGINAL_PAIRS) + len(NEW_PAIRS))
    six_pair_records = original_records + new_records
    print(f"6-pair usable items: {len(six_pair_records)}  MA-only usable items: {len(ma_records)}")

    report = {
        "_preliminary_note": (
            f"{total_pending} of 120 Massachusetts items are still pending human "
            "adjudication (MAPairs_Adjudication_2_items.xlsx) and are excluded "
            "from every number below, not estimated. Re-run once resolved."
        ),
        "pending_adjudication_ma": ma_pending,
        "six_pairs_recomputed": run_hypothesis_tests(six_pair_records, "6 PAIRS (360 sampled, recomputed here)"),
        "ma_2pairs_only": run_hypothesis_tests(ma_records, "MASSACHUSETTS 2 PAIRS ONLY (120 sampled)"),
        "pooled_8pairs": run_hypothesis_tests(six_pair_records + ma_records, "POOLED 8 PAIRS (480 sampled)"),
    }

    out_path = BASE / "full_comparison_8pairs_report.json"
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(report, f, indent=2, ensure_ascii=False)
    print(f"wrote {out_path}")


if __name__ == "__main__":
    main()
