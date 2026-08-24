"""
One-off script: builds the adjudication workbook for the 6 items where
Annotator I and Annotator J disagreed across the two new confirmatory
pairs (Tennessee Sept2024->09.11.2025, Connecticut v2024.1->v2025.1;
3 disagreements each) - reuses build_h3prime_adjudication.py's 2-rater
layout, adapted to span two pairs (an extra "pair" column, matching
run_final_metrics.py's Adjudication_43_items.xlsx column convention of
pair in column 1) so run_new_pairs_metrics.py's
_load_new_pairs_adjudication can read it directly.

Not part of the pipeline itself - a deliverable-formatting step over
already-collected annotation data.

Run:
    cd sop-guard/backend
    python -m app.research.cross_edition.annotation_packets.build_new_pairs_adjudication
"""
import sys
from pathlib import Path

BASE = Path(__file__).parent
sys.path.insert(0, str(BASE.parents[3]))  # backend/ on sys.path

from openpyxl import Workbook  # noqa: E402
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side  # noqa: E402
from openpyxl.worksheet.datavalidation import DataValidation  # noqa: E402
from openpyxl.utils import get_column_letter  # noqa: E402

from app.research.cross_edition.annotation import load_completed_xlsx, _norm_answer  # noqa: E402
from app.research.cross_edition.annotation_packets.build_annotator_workbooks import (  # noqa: E402
    load_pair, render_new_guideline, sanitize,
)

ANNOTATOR_I = BASE / "Annotator_I_ANNOTATION.xlsx"
ANNOTATOR_J = BASE / "Annotator_J_ANNOTATION.xlsx"

PAIR_SLUGS = {
    "Tennessee Sept2024→09.11.2025 (": "tennessee_sept2024_20250911",
    "Connecticut v2024.1→v2025.1 (ne": "connecticut_v20241_v20251",
}

FONT_NAME = "Calibri"
HEADER_FILL = PatternFill("solid", fgColor="C00000")
HEADER_FONT = Font(name=FONT_NAME, bold=True, color="FFFFFF", size=11)
VOTE_FILL = PatternFill("solid", fgColor="F2F2F2")
FILLIN_FILL = PatternFill("solid", fgColor="FFF2CC")
NORMAL_FONT = Font(name=FONT_NAME, size=11)
WRAP_TOP = Alignment(wrap_text=True, vertical="top")
THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
RELATION_OPTIONS = "unchanged,reworded,substantive,merged,split,moved"

COLUMNS = [
    ("pair", 22, "pair"),
    ("sample_id", 9, "sample_id"),
    ("OLD RECOMMENDATION", 40, "old_text"),
    ("NEW EDITION - full matching guideline", 50, "new_guideline_text"),
    ("Annotator I said", 25, "vote_I"),
    ("Annotator J said", 25, "vote_J"),
    ("FINAL ANSWER: correspondence", 30, "final_correspondence"),
    ("FINAL ANSWER: relation", 16, "final_relation"),
    ("Adjudication notes (why)", 30, "adj_notes"),
]
FILLIN_KEYS = {"final_correspondence", "final_relation", "adj_notes"}

INSTRUCTIONS = [
    ("What this sheet is", None),
    (None, "Two people (I and J) independently labelled the same two new "
           "60-item packets (Tennessee and Connecticut). On these 6 items "
           "(3 per pair), their answers disagreed."),
    ("What to do", None),
    (None, "For each row: read the old recommendation (column C) and the "
           "new edition's guideline (column D). Look at what both "
           "annotators said (columns E-F). Decide what the ACTUAL correct "
           "answer is."),
    (None, "Then fill in columns G-I (yellow): the final correspondence "
           "(item ID, NONE, or CANNOT_DETERMINE), the final relation, and a "
           "short note on why."),
    ("Important", None),
    (None, "If you genuinely cannot resolve one, CANNOT_DETERMINE as the "
           "final answer is a legitimate outcome, not a failure to "
           "adjudicate."),
]


def vote_text(raw_entry: dict) -> str:
    corr = raw_entry.get("correspondence") or ""
    rel = raw_entry.get("relation") or ""
    return f"{corr}" + (f"  ({rel})" if rel else "")


def build_instructions_sheet(wb: Workbook):
    ws = wb.active
    ws.title = "READ ME FIRST"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 100
    r = 1
    c = ws.cell(row=r, column=1, value="New-pairs adjudication - 6 disputed items (I/J)")
    c.font = Font(name=FONT_NAME, bold=True, size=16, color="C00000")
    r += 2
    for heading, body in INSTRUCTIONS:
        if heading is not None:
            c = ws.cell(row=r, column=1, value=heading)
            c.font = Font(name=FONT_NAME, bold=True, size=13, color="1F4E78")
            r += 1
        if body is not None:
            c = ws.cell(row=r, column=1, value=body)
            c.font = Font(name=FONT_NAME, size=11.5)
            c.alignment = Alignment(wrap_text=True, vertical="top")
            ws.row_dimensions[r].height = 34
            r += 1


def build_data_sheet(wb: Workbook, disputes: list[tuple[str, str]], raw_i: dict, raw_j: dict):
    """disputes: list of (pair_sheet_title, sample_id)."""
    ws = wb.create_sheet(title="Adjudicate these 6")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"

    for col_idx, (header, width, _key) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = WRAP_TOP
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[1].height = 30

    dv = DataValidation(type="list", formula1=f'"{RELATION_OPTIONS}"', allow_blank=True)
    dv.error = "Pick one from the list."
    dv.errorTitle = "Invalid entry"
    ws.add_data_validation(dv)

    packets = {}
    for pair in PAIR_SLUGS:
        rows, ctx = load_pair(PAIR_SLUGS[pair])
        packet = {row["sample_id"]: row for row in rows}
        packets[pair] = (packet, ctx)

    r = 2
    for pair, sid in disputes:
        packet, ctx = packets[pair]
        src = packet[sid]
        values = {
            "pair": pair,
            "sample_id": sid,
            "old_text": sanitize(src["old_text"]),
            "new_guideline_text": render_new_guideline(ctx.get(sid, {})),
            "vote_I": sanitize(vote_text(raw_i[pair][sid])),
            "vote_J": sanitize(vote_text(raw_j[pair][sid])),
            "final_correspondence": "",
            "final_relation": "",
            "adj_notes": "",
        }
        for col_idx, (_h, _w, key) in enumerate(COLUMNS, start=1):
            cell = ws.cell(row=r, column=col_idx, value=values[key])
            cell.border = BORDER
            cell.alignment = WRAP_TOP
            if key in FILLIN_KEYS:
                cell.fill = FILLIN_FILL
            elif key.startswith("vote_"):
                cell.fill = VOTE_FILL
                cell.font = NORMAL_FONT
            else:
                cell.font = NORMAL_FONT
        ws.row_dimensions[r].height = 60
        dv.add(ws.cell(row=r, column=8))  # column H = final_relation
        r += 1

    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{r - 1}"
    return r - 2


if __name__ == "__main__":
    raw_i = load_completed_xlsx(str(ANNOTATOR_I))
    raw_j = load_completed_xlsx(str(ANNOTATOR_J))

    disputes: list[tuple[str, str]] = []
    for pair in PAIR_SLUGS:
        ai = {sid: _norm_answer(v["correspondence"]) for sid, v in raw_i[pair].items()}
        aj = {sid: _norm_answer(v["correspondence"]) for sid, v in raw_j[pair].items()}
        for sid in sorted(ai):
            if ai.get(sid) != aj.get(sid):
                disputes.append((pair, sid))
    print(f"disputed items: {disputes}")

    wb = Workbook()
    build_instructions_sheet(wb)
    n = build_data_sheet(wb, disputes, raw_i, raw_j)

    out_path = BASE / "NewPairs_Adjudication_6_items.xlsx"
    wb.save(out_path)
    print(f"wrote {out_path} ({n} rows)")
